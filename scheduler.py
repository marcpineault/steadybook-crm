"""
Scheduler module for Calm Money Pipeline Bot.

Provides:
- Morning briefing at 8:00 AM ET daily
- Auto-nag system every 2 hours (9AM-5PM ET)

Usage:
    from scheduler import start_scheduler
    start_scheduler(telegram_app)
"""

import json
import logging
import os
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pytz
from apscheduler.schedulers.asyncio import AsyncIOScheduler

logger = logging.getLogger(__name__)

DATA_DIR = os.environ.get("DATA_DIR", "")
if DATA_DIR:
    PIPELINE_PATH = os.path.join(DATA_DIR, "pipeline.xlsx")
else:
    PIPELINE_PATH = os.environ.get("PIPELINE_PATH", "pipeline.xlsx")
CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
if DATA_DIR:
    NAG_STATE_FILE = os.path.join(DATA_DIR, "nag_state.json")
else:
    NAG_STATE_FILE = "nag_state.json"
ET = pytz.timezone("America/Toronto")

# Pipeline sheet layout
PIPELINE_DATA_START = 5
PIPELINE_MAX_ROWS = 80
PIPELINE_COLS = {
    "name": 1, "phone": 2, "email": 3, "source": 4,
    "priority": 5, "stage": 6, "product": 7,
    "aum": 8, "revenue": 9, "first_contact": 10,
    "next_followup": 11, "days_open": 12, "notes": 13,
}

# Meetings sheet layout
MEETINGS_DATA_START = 3
MEETINGS_COLS = {
    "date": 1, "time": 2, "prospect": 3,
    "type": 4, "prep_notes": 5, "status": 6,
}

# Insurance Book sheet layout
INSURANCE_DATA_START = 3
INSURANCE_COLS = {
    "name": 1, "phone": 2, "address": 3, "policy_start": 4,
    "status": 5, "last_called": 6, "notes": 7, "retry_date": 8,
}

INSURANCE_DAILY_LIMIT = 5

# ── Bot reference ──

_bot = None


# ── Nag state persistence ──

def _load_nag_state() -> dict:
    if Path(NAG_STATE_FILE).exists():
        try:
            with open(NAG_STATE_FILE, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def _save_nag_state(state: dict):
    with open(NAG_STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)


def _can_nag(state: dict, prospect_name: str, nag_type: str) -> bool:
    """Check if we can nag about this prospect (not within 24 hours of last nag)."""
    key = f"{prospect_name}::{nag_type}"
    last = state.get(key)
    if not last:
        return True
    try:
        last_dt = datetime.fromisoformat(last)
        return datetime.now() - last_dt > timedelta(hours=24)
    except (ValueError, TypeError):
        return True


def _mark_nagged(state: dict, prospect_name: str, nag_type: str):
    key = f"{prospect_name}::{nag_type}"
    state[key] = datetime.now().isoformat()


# ── Excel readers ──

def _parse_date(val) -> date | None:
    """Parse a date value from Excel (could be datetime, date, or string)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip().split(" ")[0], "%Y-%m-%d").date()
    except (ValueError, IndexError):
        pass
    # Try other common formats
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _cell(ws, row, col):
    """Get cell value, return empty string for None."""
    v = ws.cell(row=row, column=col).value
    return str(v) if v is not None else ""


def _get_lock():
    from bot import pipeline_lock
    return pipeline_lock


def _read_prospects():
    """Read pipeline prospects."""
    with _get_lock():
        return _read_prospects_inner()


def _read_prospects_inner():
    wb = openpyxl.load_workbook(PIPELINE_PATH, data_only=True)
    ws = wb["Pipeline"]
    prospects = []
    for r in range(PIPELINE_DATA_START, PIPELINE_DATA_START + PIPELINE_MAX_ROWS):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        p = {"row": r}
        for field, col in PIPELINE_COLS.items():
            p[field] = _cell(ws, r, col)
        p["_next_followup_date"] = _parse_date(ws.cell(row=r, column=PIPELINE_COLS["next_followup"]).value)
        p["_first_contact_date"] = _parse_date(ws.cell(row=r, column=PIPELINE_COLS["first_contact"]).value)
        prospects.append(p)
    wb.close()
    return prospects


def _read_meetings_today():
    """Read meetings for today from the Meetings sheet."""
    with _get_lock():
        return _read_meetings_today_inner()


def _read_meetings_today_inner():
    wb = openpyxl.load_workbook(PIPELINE_PATH, data_only=True)
    if "Meetings" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Meetings"]
    today = date.today()
    meetings = []
    for r in range(MEETINGS_DATA_START, MEETINGS_DATA_START + 100):
        raw_date = ws.cell(row=r, column=MEETINGS_COLS["date"]).value
        if not raw_date:
            continue
        meeting_date = _parse_date(raw_date)
        if meeting_date == today:
            meetings.append({
                "time": _cell(ws, r, MEETINGS_COLS["time"]),
                "prospect": _cell(ws, r, MEETINGS_COLS["prospect"]),
                "type": _cell(ws, r, MEETINGS_COLS["type"]),
                "prep_notes": _cell(ws, r, MEETINGS_COLS["prep_notes"]),
                "status": _cell(ws, r, MEETINGS_COLS["status"]),
            })
    wb.close()
    return meetings


def _read_meetings_tomorrow():
    """Read meetings for tomorrow from the Meetings sheet."""
    with _get_lock():
        return _read_meetings_tomorrow_inner()


def _read_meetings_tomorrow_inner():
    wb = openpyxl.load_workbook(PIPELINE_PATH, data_only=True)
    if "Meetings" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Meetings"]
    tomorrow = date.today() + timedelta(days=1)
    meetings = []
    for r in range(MEETINGS_DATA_START, MEETINGS_DATA_START + 100):
        raw_date = ws.cell(row=r, column=MEETINGS_COLS["date"]).value
        if not raw_date:
            continue
        meeting_date = _parse_date(raw_date)
        if meeting_date == tomorrow:
            meetings.append({
                "time": _cell(ws, r, MEETINGS_COLS["time"]),
                "prospect": _cell(ws, r, MEETINGS_COLS["prospect"]),
                "type": _cell(ws, r, MEETINGS_COLS["type"]),
                "prep_notes": _cell(ws, r, MEETINGS_COLS["prep_notes"]),
                "status": _cell(ws, r, MEETINGS_COLS["status"]),
            })
    wb.close()
    return meetings


def _read_insurance_calls():
    """Read insurance book entries that need calling today."""
    with _get_lock():
        return _read_insurance_calls_inner()


def _read_insurance_calls_inner():
    wb = openpyxl.load_workbook(PIPELINE_PATH, data_only=True)
    if "Insurance Book" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Insurance Book"]
    today = date.today()
    calls = []
    for r in range(INSURANCE_DATA_START, INSURANCE_DATA_START + 200):
        name = ws.cell(row=r, column=INSURANCE_COLS["name"]).value
        if not name:
            continue
        status = _cell(ws, r, INSURANCE_COLS["status"]).strip()
        retry_raw = ws.cell(row=r, column=INSURANCE_COLS["retry_date"]).value
        retry_date = _parse_date(retry_raw)

        eligible = False
        if status.lower() in ("not called", ""):
            eligible = True
        elif retry_date and retry_date <= today:
            eligible = True

        if eligible:
            calls.append({
                "name": _cell(ws, r, INSURANCE_COLS["name"]),
                "phone": _cell(ws, r, INSURANCE_COLS["phone"]),
                "address": _cell(ws, r, INSURANCE_COLS["address"]),
                "policy_start": _cell(ws, r, INSURANCE_COLS["policy_start"]),
                "status": status,
                "notes": _cell(ws, r, INSURANCE_COLS["notes"]),
            })
            if len(calls) >= INSURANCE_DAILY_LIMIT:
                break
    wb.close()
    return calls


# ── Morning Briefing ──

async def morning_briefing():
    """Send the daily morning briefing at 8:00 AM ET."""
    if not _bot or not CHAT_ID:
        logger.warning("Bot or CHAT_ID not configured, skipping morning briefing.")
        return

    if not Path(PIPELINE_PATH).exists():
        await _bot.send_message(chat_id=CHAT_ID, text="Good morning! Pipeline file not found — upload one to get started.")
        return

    today = date.today()
    lines = [f"Good morning, Marc! Here's your briefing for {today.strftime('%A, %B %d')}.\n"]

    try:
        prospects = _read_prospects()
    except Exception as e:
        logger.error(f"Error reading pipeline: {e}")
        await _bot.send_message(chat_id=CHAT_ID, text=f"Morning briefing error reading pipeline: {e}")
        return

    active = [p for p in prospects if p["stage"] not in ("Closed-Won", "Closed-Lost", "")]

    # ── Follow-ups due today ──
    due_today = []
    for p in active:
        fu = p["_next_followup_date"]
        if fu == today:
            notes_snippet = p["notes"][:80] + "..." if len(p["notes"]) > 80 else p["notes"]
            due_today.append(f"  - {p['name']} [{p['priority']}] — {notes_snippet or 'no notes'}")

    if due_today:
        lines.append(f"FOLLOW-UPS DUE TODAY ({len(due_today)}):")
        lines.extend(due_today)
        lines.append("")

    # ── Overdue follow-ups ──
    overdue = []
    for p in active:
        fu = p["_next_followup_date"]
        if fu and fu < today:
            days_late = (today - fu).days
            overdue.append((days_late, f"  - {p['name']} — {days_late} day{'s' if days_late != 1 else ''} late"))

    overdue.sort(key=lambda x: -x[0])  # most overdue first
    if overdue:
        lines.append(f"OVERDUE ({len(overdue)}):")
        lines.extend([item[1] for item in overdue])
        lines.append("")

    # ── Meetings today ──
    try:
        meetings = _read_meetings_today()
        if meetings:
            lines.append(f"MEETINGS TODAY ({len(meetings)}):")
            for m in meetings:
                status_tag = f" [{m['status']}]" if m["status"] else ""
                prep = f" | Prep: {m['prep_notes']}" if m["prep_notes"] else ""
                lines.append(f"  - {m['time']} — {m['prospect']} ({m['type']}){status_tag}{prep}")
            lines.append("")
    except Exception as e:
        logger.warning(f"Could not read Meetings sheet: {e}")

    # ── Insurance book calls ──
    try:
        calls = _read_insurance_calls()
        if calls:
            lines.append(f"INSURANCE BOOK CALLS ({len(calls)}):")
            for c in calls:
                notes_bit = f" | {c['notes']}" if c["notes"] else ""
                lines.append(f"  - {c['name']} — {c['phone']}{notes_bit}")
            lines.append("")
    except Exception as e:
        logger.warning(f"Could not read Insurance Book sheet: {e}")

    # ── Pipeline snapshot ──
    total_aum = 0
    for p in active:
        try:
            total_aum += float(p["aum"].replace("$", "").replace(",", "")) if p["aum"] else 0
        except ValueError:
            pass
    hot_count = len([p for p in active if p["priority"].lower() == "hot"])

    lines.append("PIPELINE SNAPSHOT:")
    lines.append(f"  Active: {len(active)} | Value: ${total_aum:,.0f} | Hot: {hot_count}")

    if not due_today and not overdue:
        lines.append("\nCalendar is clear of follow-ups. Good day to prospect!")

    msg = "\n".join(lines)
    await _bot.send_message(chat_id=CHAT_ID, text=msg)
    logger.info("Morning briefing sent.")


# ── Auto-Nag System ──

async def auto_nag():
    """Check for items that need attention and send nag messages."""
    if not _bot or not CHAT_ID:
        logger.warning("Bot or CHAT_ID not configured, skipping auto-nag.")
        return

    if not Path(PIPELINE_PATH).exists():
        return

    today = date.today()
    nag_state = _load_nag_state()
    alerts = []

    try:
        prospects = _read_prospects()
    except Exception as e:
        logger.error(f"Error reading pipeline for nag: {e}")
        return

    active = [p for p in prospects if p["stage"] not in ("Closed-Won", "Closed-Lost", "")]

    for p in active:
        name = p["name"]
        fu = p["_next_followup_date"]
        fc = p["_first_contact_date"]
        stage = p["stage"]
        priority = p["priority"].lower()

        # 1. No activity for 7+ days
        ref_date = fu or fc
        if ref_date and (today - ref_date).days >= 7:
            if _can_nag(nag_state, name, "stale"):
                days_idle = (today - ref_date).days
                alerts.append(f"  STALE: {name} — no activity for {days_idle} days")
                _mark_nagged(nag_state, name, "stale")

        # 2. Follow-up 2+ days overdue
        if fu and (today - fu).days >= 2:
            if _can_nag(nag_state, name, "overdue"):
                days_late = (today - fu).days
                alerts.append(f"  OVERDUE: {name} — follow-up is {days_late} days late")
                _mark_nagged(nag_state, name, "overdue")

        # 3. Hot lead stuck in New Lead/Contacted for 5+ days
        if priority == "hot" and stage in ("New Lead", "Contacted"):
            if fc and (today - fc).days >= 5:
                if _can_nag(nag_state, name, "hot_stuck"):
                    alerts.append(f"  HOT STUCK: {name} — hot lead in '{stage}' for {(today - fc).days} days")
                    _mark_nagged(nag_state, name, "hot_stuck")

    # 4. Meeting tomorrow prep reminder
    try:
        tomorrow_meetings = _read_meetings_tomorrow()
        for m in tomorrow_meetings:
            mkey = f"{m['prospect']}_{m['time']}"
            if _can_nag(nag_state, mkey, "meeting_prep"):
                prep = f" — Prep: {m['prep_notes']}" if m["prep_notes"] else ""
                alerts.append(f"  MEETING TOMORROW: {m['prospect']} at {m['time']} ({m['type']}){prep}")
                _mark_nagged(nag_state, mkey, "meeting_prep")
    except Exception as e:
        logger.warning(f"Could not check tomorrow's meetings for nag: {e}")

    _save_nag_state(nag_state)

    if alerts:
        header = f"Hey Marc, heads up ({len(alerts)} item{'s' if len(alerts) != 1 else ''}):\n"
        msg = header + "\n".join(alerts)
        await _bot.send_message(chat_id=CHAT_ID, text=msg)
        logger.info(f"Auto-nag sent with {len(alerts)} alerts.")


# ── Weekly Performance Report ──

async def weekly_report():
    """Send weekly performance report Sunday at 7 PM ET."""
    if not _bot or not CHAT_ID:
        return

    if not Path(PIPELINE_PATH).exists():
        return

    today = date.today()
    week_start = today - timedelta(days=7)

    try:
        prospects = _read_prospects()
    except Exception as e:
        logger.error(f"Error reading pipeline for weekly report: {e}")
        return

    active = [p for p in prospects if p["stage"] not in ("Closed-Won", "Closed-Lost", "")]
    won = [p for p in prospects if p["stage"] == "Closed-Won"]

    # Pipeline value
    total_aum = 0
    total_rev = 0
    for p in active:
        try:
            total_aum += float(p["aum"].replace("$", "").replace(",", "")) if p["aum"] else 0
        except ValueError:
            pass
        try:
            total_rev += float(p["revenue"].replace("$", "").replace(",", "")) if p["revenue"] else 0
        except ValueError:
            pass

    won_rev = 0
    for p in won:
        try:
            won_rev += float(p["revenue"].replace("$", "").replace(",", "")) if p["revenue"] else 0
        except ValueError:
            pass

    hot_count = len([p for p in active if p["priority"].lower() == "hot"])

    # Activity log this week
    with _get_lock():
        wb = openpyxl.load_workbook(PIPELINE_PATH, data_only=True)
        week_activities = 0
        calls_made = 0
        emails_sent = 0
        meetings_held = 0
        if "Activity Log" in wb.sheetnames:
            log_ws = wb["Activity Log"]
            for r in range(3, 200):
                d = log_ws.cell(row=r, column=1).value
                if not d:
                    continue
                activity_date = _parse_date(d)
                if activity_date and activity_date >= week_start:
                    week_activities += 1
                    action = _cell(log_ws, r, 3).lower()
                    if "call" in action or "phone" in action:
                        calls_made += 1
                    if "email" in action:
                        emails_sent += 1
                    if "meeting" in action or "discovery" in action or "presentation" in action:
                        meetings_held += 1

        # Insurance book stats this week
        book_calls_week = 0
        book_booked = 0
        if "Insurance Book" in wb.sheetnames:
            bs = wb["Insurance Book"]
            for r in range(INSURANCE_DATA_START, INSURANCE_DATA_START + 200):
                name = bs.cell(row=r, column=INSURANCE_COLS["name"]).value
                if not name:
                    continue
                last_called = _parse_date(bs.cell(row=r, column=INSURANCE_COLS["last_called"]).value)
                if last_called and last_called >= week_start:
                    book_calls_week += 1
                status = _cell(bs, r, INSURANCE_COLS["status"]).lower()
                if status == "booked meeting":
                    lc = _parse_date(bs.cell(row=r, column=INSURANCE_COLS["last_called"]).value)
                    if lc and lc >= week_start:
                        book_booked += 1

        # Win/loss this week
        wins_week = 0
        losses_week = 0
        if "Win Loss Log" in wb.sheetnames:
            wl = wb["Win Loss Log"]
            for r in range(3, 103):
                d = wl.cell(row=r, column=1).value
                if not d:
                    continue
                wl_date = _parse_date(d)
                outcome = _cell(wl, r, 3).lower()
                if wl_date and wl_date >= week_start:
                    if outcome in ("won", "closed-won"):
                        wins_week += 1
                    elif outcome in ("lost", "closed-lost"):
                        losses_week += 1

        wb.close()

    # Overdue count
    overdue_count = 0
    for p in active:
        fu = p["_next_followup_date"]
        if fu and fu < today:
            overdue_count += 1

    # Build the report
    lines = [
        f"WEEKLY REPORT — {week_start.strftime('%b %d')} to {today.strftime('%b %d')}",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        "",
        "ACTIVITY:",
        f"  Total touchpoints: {week_activities}",
        f"  Calls: {calls_made} | Emails: {emails_sent} | Meetings: {meetings_held}",
    ]

    if book_calls_week > 0:
        lines.append(f"  Book calls: {book_calls_week} | Booked: {book_booked}")

    lines.extend([
        "",
        "RESULTS:",
        f"  Deals won: {wins_week} | Lost: {losses_week}",
        f"  Won revenue: ${won_rev:,.0f}",
        "",
        "PIPELINE:",
        f"  Active: {len(active)} | Value: ${total_aum:,.0f} | Hot: {hot_count}",
        f"  Est. revenue: ${total_rev:,.0f}",
        f"  Overdue follow-ups: {overdue_count}",
        "",
        "MONDAY PRIORITIES:",
    ])

    # Top 3 priorities for Monday
    priorities = []
    # Overdue hot leads first
    for p in active:
        fu = p["_next_followup_date"]
        if p["priority"].lower() == "hot" and fu and fu < today:
            priorities.append(f"  1. Follow up with {p['name']} (Hot, {(today - fu).days}d overdue)")
    # Then other overdue
    for p in active:
        fu = p["_next_followup_date"]
        if p["priority"].lower() != "hot" and fu and fu < today:
            priorities.append(f"  {len(priorities)+1}. Follow up with {p['name']} ({(today - fu).days}d overdue)")
        if len(priorities) >= 3:
            break

    if not priorities:
        priorities.append("  All caught up! Focus on prospecting.")

    lines.extend(priorities[:3])
    lines.append(f"\nKeep grinding, Marc. 💪")

    msg = "\n".join(lines)
    await _bot.send_message(chat_id=CHAT_ID, text=msg)
    logger.info("Weekly report sent.")


# ── Scheduler entry point ──

def start_scheduler(telegram_app):
    """
    Start the scheduler with the Telegram application.

    Args:
        telegram_app: The python-telegram-bot Application object.
    """
    global _bot
    _bot = telegram_app.bot

    if not CHAT_ID:
        logger.warning("TELEGRAM_CHAT_ID not set — scheduler will not send messages.")
        return

    scheduler = AsyncIOScheduler(timezone=ET)

    # Morning briefing at 8:00 AM ET every day
    scheduler.add_job(
        morning_briefing,
        "cron",
        hour=8,
        minute=0,
        id="morning_briefing",
        name="Daily Morning Briefing",
    )

    # Auto-nag every 2 hours from 9 AM to 5 PM ET
    scheduler.add_job(
        auto_nag,
        "cron",
        hour="9,11,13,15,17",
        minute=0,
        id="auto_nag",
        name="Auto-Nag Check",
    )

    # Weekly performance report Sunday at 7 PM ET
    scheduler.add_job(
        weekly_report,
        "cron",
        day_of_week="sun",
        hour=19,
        minute=0,
        id="weekly_report",
        name="Weekly Performance Report",
    )

    scheduler.start()
    logger.info("Scheduler started — morning briefing 8AM, auto-nag 9-5, weekly report Sun 7PM ET.")
