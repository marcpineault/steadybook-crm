import html as _html
import os
import re
import threading
from datetime import date, datetime, timedelta
from pathlib import Path

import json

import openpyxl
from flask import Flask, Response, request, jsonify


def _esc(val):
    """Escape HTML to prevent XSS."""
    return _html.escape(str(val)) if val else ""

DATA_DIR = os.environ.get("DATA_DIR", "")
if DATA_DIR:
    PIPELINE_PATH = os.path.join(DATA_DIR, "pipeline.xlsx")
else:
    PIPELINE_PATH = os.environ.get("PIPELINE_PATH", "pipeline.xlsx")

app = Flask(__name__)

DATA_START = 5
MAX_ROWS = 80

STAGE_COLORS = {
    "New Lead": "#BDC3C7", "Contacted": "#3498DB", "Discovery Call": "#8E44AD",
    "Needs Analysis": "#E67E22", "Plan Presentation": "#F39C12", "Proposal Sent": "#2980B9",
    "Negotiation": "#E74C3C", "Closed-Won": "#27AE60", "Closed-Lost": "#95A5A6", "Nurture": "#1ABC9C",
}
PRIORITY_COLORS = {"Hot": "#E74C3C", "Warm": "#F39C12", "Cold": "#3498DB"}
STAGE_TEXT = {
    "New Lead": "#2C3E50", "Plan Presentation": "#2C3E50",
}


def read_data():
    if not Path(PIPELINE_PATH).exists():
        return [], [], [], []

    lock = _get_lock()
    lock.acquire()
    try:
        return _read_data_inner()
    finally:
        lock.release()


def _read_data_inner():
    wb = openpyxl.load_workbook(PIPELINE_PATH, data_only=True)

    # Pipeline
    ws = wb["Pipeline"]
    prospects = []
    for r in range(DATA_START, DATA_START + MAX_ROWS):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        prospects.append({
            "name": str(name),
            "phone": str(ws.cell(row=r, column=2).value or ""),
            "email": str(ws.cell(row=r, column=3).value or ""),
            "source": str(ws.cell(row=r, column=4).value or ""),
            "priority": str(ws.cell(row=r, column=5).value or ""),
            "stage": str(ws.cell(row=r, column=6).value or ""),
            "product": str(ws.cell(row=r, column=7).value or ""),
            "aum": ws.cell(row=r, column=8).value or 0,
            "revenue": ws.cell(row=r, column=9).value or 0,
            "first_contact": str(ws.cell(row=r, column=10).value or ""),
            "next_followup": str(ws.cell(row=r, column=11).value or ""),
            "notes": str(ws.cell(row=r, column=13).value or ""),
        })

    # Activity log
    log_ws = wb["Activity Log"]
    activities = []
    for r in range(3, 103):
        d = log_ws.cell(row=r, column=1).value
        if not d:
            continue
        activities.append({
            "date": str(d),
            "prospect": str(log_ws.cell(row=r, column=2).value or ""),
            "action": str(log_ws.cell(row=r, column=3).value or ""),
            "outcome": str(log_ws.cell(row=r, column=4).value or ""),
            "next_step": str(log_ws.cell(row=r, column=5).value or ""),
        })

    # Meetings
    meetings = []
    if "Meetings" in wb.sheetnames:
        ms = wb["Meetings"]
        for r in range(3, 103):
            d = ms.cell(row=r, column=1).value
            if not d:
                continue
            meetings.append({
                "date": str(d).split(" ")[0] if d else "",
                "time": str(ms.cell(row=r, column=2).value or ""),
                "prospect": str(ms.cell(row=r, column=3).value or ""),
                "type": str(ms.cell(row=r, column=4).value or ""),
                "prep_notes": str(ms.cell(row=r, column=5).value or ""),
                "status": str(ms.cell(row=r, column=6).value or "Scheduled"),
            })

    # Insurance Book
    book_entries = []
    if "Insurance Book" in wb.sheetnames:
        bs = wb["Insurance Book"]
        for r in range(3, 203):
            name = bs.cell(row=r, column=1).value
            if not name:
                continue
            book_entries.append({
                "name": str(name),
                "phone": str(bs.cell(row=r, column=2).value or ""),
                "address": str(bs.cell(row=r, column=3).value or ""),
                "policy_start": str(bs.cell(row=r, column=4).value or ""),
                "status": str(bs.cell(row=r, column=5).value or "Not Called"),
                "last_called": str(bs.cell(row=r, column=6).value or ""),
                "notes": str(bs.cell(row=r, column=7).value or ""),
                "retry_date": str(bs.cell(row=r, column=8).value or ""),
            })

    wb.close()
    return prospects, activities, meetings, book_entries


PIPELINE_COLS = {
    "name": 1, "phone": 2, "email": 3, "source": 4,
    "priority": 5, "stage": 6, "product": 7,
    "aum": 8, "revenue": 9, "first_contact": 10,
    "next_followup": 11, "notes": 13,
}


def _get_lock():
    from bot import pipeline_lock
    return pipeline_lock


@app.route("/api/prospect", methods=["POST"])
def api_add_prospect():
    data = request.json
    if not data or not data.get("name"):
        return jsonify({"error": "Name required"}), 400

    with _get_lock():
        wb = openpyxl.load_workbook(PIPELINE_PATH)
        ws = wb["Pipeline"]

        target_row = None
        for r in range(DATA_START, DATA_START + MAX_ROWS):
            if not ws.cell(row=r, column=1).value:
                target_row = r
                break

        if not target_row:
            wb.close()
            return jsonify({"error": "Pipeline full"}), 400

        for field, col in PIPELINE_COLS.items():
            val = data.get(field, "")
            if val:
                ws.cell(row=target_row, column=col, value=val)

        if not data.get("first_contact"):
            ws.cell(row=target_row, column=10, value=date.today().strftime("%Y-%m-%d"))
        if not data.get("stage"):
            ws.cell(row=target_row, column=6, value="New Lead")

        wb.save(PIPELINE_PATH)
        wb.close()
    return jsonify({"ok": True, "row": target_row})


@app.route("/api/prospect/<name>", methods=["PUT"])
def api_update_prospect(name):
    data = request.json
    if not data:
        return jsonify({"error": "No data"}), 400

    with _get_lock():
        wb = openpyxl.load_workbook(PIPELINE_PATH)
        ws = wb["Pipeline"]

        found_row = None
        for r in range(DATA_START, DATA_START + MAX_ROWS):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val and str(cell_val).strip().lower() == name.strip().lower():
                found_row = r
                break

        if not found_row:
            wb.close()
            return jsonify({"error": f"Prospect '{name}' not found"}), 404

        for field, col in PIPELINE_COLS.items():
            if field in data:
                ws.cell(row=found_row, column=col, value=data[field])

        wb.save(PIPELINE_PATH)
        wb.close()
    return jsonify({"ok": True})


@app.route("/api/prospect/<name>", methods=["DELETE"])
def api_delete_prospect(name):
    with _get_lock():
        wb = openpyxl.load_workbook(PIPELINE_PATH)
        ws = wb["Pipeline"]

        found_row = None
        for r in range(DATA_START, DATA_START + MAX_ROWS):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val and str(cell_val).strip().lower() == name.strip().lower():
                found_row = r
                break

        if not found_row:
            wb.close()
            return jsonify({"error": f"Prospect '{name}' not found"}), 404

        for col in range(1, 14):
            ws.cell(row=found_row, column=col, value=None)

        wb.save(PIPELINE_PATH)
        wb.close()
    return jsonify({"ok": True})


@app.route("/api/prospects")
def api_list_prospects():
    prospects, _, _, _ = read_data()
    return jsonify(prospects)


def calc_fyc(premium, product):
    """Calculate First Year Commission from premium and product term.
    Term 20/25/30: Premium * 11.11 * 0.5
    Term 10/15:    Premium * 11.11 * 0.4
    """
    prem = parse_money(premium) if not isinstance(premium, (int, float)) else premium
    if prem <= 0:
        return 0.0
    # Extract term number from product string like "Term 20", "Versatile Term 30", etc.
    term_match = re.search(r'(\d+)', str(product or ""))
    if not term_match:
        return 0.0
    term = int(term_match.group(1))
    if term in (20, 25, 30):
        return prem * 11.11 * 0.5
    elif term in (10, 15):
        return prem * 11.11 * 0.4
    return 0.0


def parse_money(val):
    """Parse money values like '200k', '1.5M', '$500,000' into float."""
    try:
        s = str(val).strip().replace("$", "").replace(",", "").lower()
        if not s or s == "none" or s == "0":
            return 0.0
        if s.endswith("m"):
            return float(s[:-1]) * 1000000
        if s.endswith("k"):
            return float(s[:-1]) * 1000
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def fmt_money(val):
    try:
        v = float(str(val).replace("$", "").replace(",", ""))
        if v >= 1000000:
            return f"${v/1000000:.1f}M"
        if v >= 1000:
            return f"${v/1000:.0f}K"
        return f"${v:,.0f}"
    except (ValueError, TypeError):
        return "$0"


def fmt_money_full(val):
    try:
        return f"${float(str(val).replace('$','').replace(',','')):,.0f}"
    except (ValueError, TypeError):
        return "$0"


@app.route("/")
def dashboard():
    prospects, activities, meetings, book_entries = read_data()
    today = date.today()
    now = datetime.now()

    active = [p for p in prospects if p["stage"] not in ("Closed-Won", "Closed-Lost", "")]
    won = [p for p in prospects if p["stage"] == "Closed-Won"]
    lost = [p for p in prospects if p["stage"] == "Closed-Lost"]

    total_pipeline = sum(parse_money(p["aum"]) for p in active)
    total_revenue = sum(parse_money(p["revenue"]) for p in active)
    won_revenue = sum(parse_money(p["revenue"]) for p in won)
    hot_count = len([p for p in active if p["priority"] == "Hot"])

    win_rate = 0
    if len(won) + len(lost) > 0:
        win_rate = len(won) / (len(won) + len(lost)) * 100

    overdue = []
    for p in active:
        if p["next_followup"] and p["next_followup"] != "None":
            try:
                fu = datetime.strptime(p["next_followup"].split(" ")[0], "%Y-%m-%d").date()
                if fu < today:
                    overdue.append(p)
            except (ValueError, IndexError):
                pass

    # ── Revenue Forecasting ──
    PREMIUM_TARGET = 200000
    AUM_TARGET = 5000000

    # Baseline numbers (existing business not tracked in pipeline)
    # These ONLY apply to forecast targets, not main KPI cards
    BASELINE_AUM = 400000    # Current AUM as of March 2025
    BASELINE_PREMIUM = 2000  # Life premium YTD

    # Calculate won AUM from pipeline only (for KPI cards)
    won_aum_pipeline = 0
    for p in won:
        try:
            won_aum_pipeline += parse_money(p["aum"])
        except ValueError:
            pass

    # FYC calculations
    won_fyc = sum(calc_fyc(p["revenue"], p["product"]) for p in won)
    active_fyc = sum(calc_fyc(p["revenue"], p["product"]) for p in active)

    # Forecast totals include baselines
    forecast_revenue = won_revenue + BASELINE_PREMIUM
    forecast_aum = won_aum_pipeline + BASELINE_AUM

    # Days into the year / days remaining
    year_start = date(today.year, 1, 1)
    year_end = date(today.year, 12, 31)
    days_elapsed = (today - year_start).days + 1
    days_total = (year_end - year_start).days + 1
    days_remaining = days_total - days_elapsed
    pct_year = days_elapsed / days_total * 100

    # Premium progress (forecast includes baseline)
    premium_pct = (forecast_revenue / PREMIUM_TARGET * 100) if PREMIUM_TARGET else 0
    premium_pace = (forecast_revenue / days_elapsed * days_total) if days_elapsed else 0
    premium_on_pace = premium_pace >= PREMIUM_TARGET

    # AUM progress (forecast includes baseline)
    aum_pct = (forecast_aum / AUM_TARGET * 100) if AUM_TARGET else 0
    aum_pace = (forecast_aum / days_elapsed * days_total) if days_elapsed else 0
    aum_on_pace = aum_pace >= AUM_TARGET

    # Pipeline weighted forecast (probability by stage)
    stage_probability = {
        "New Lead": 0.05, "Contacted": 0.10, "Discovery Call": 0.20,
        "Needs Analysis": 0.35, "Plan Presentation": 0.50, "Proposal Sent": 0.65,
        "Negotiation": 0.80, "Nurture": 0.05,
    }
    weighted_revenue = 0
    weighted_aum = 0
    weighted_fyc = 0
    for p in active:
        prob = stage_probability.get(p["stage"], 0.10)
        try:
            weighted_revenue += parse_money(p["revenue"]) * prob
        except ValueError:
            pass
        try:
            weighted_aum += parse_money(p["aum"]) * prob
        except ValueError:
            pass
        weighted_fyc += calc_fyc(p["revenue"], p["product"]) * prob

    projected_revenue = forecast_revenue + weighted_revenue
    projected_aum = forecast_aum + weighted_aum
    projected_fyc = won_fyc + weighted_fyc

    # Monthly revenue tracking (won deals by month)
    monthly_revenue = {}
    monthly_aum = {}
    for p in won:
        fc = p.get("first_contact", "")
        if fc and fc != "None":
            try:
                m = datetime.strptime(fc.split(" ")[0], "%Y-%m-%d")
                if m.year == today.year:
                    month_key = m.strftime("%b")
                    month_num = m.month
                    try:
                        monthly_revenue[(month_num, month_key)] = monthly_revenue.get((month_num, month_key), 0) + parse_money(p["revenue"])
                    except ValueError:
                        pass
                    try:
                        monthly_aum[(month_num, month_key)] = monthly_aum.get((month_num, month_key), 0) + parse_money(p["aum"])
                    except ValueError:
                        pass
            except (ValueError, IndexError):
                pass

    # Fill all months up to current
    all_months = []
    for m in range(1, today.month + 1):
        mk = date(today.year, m, 1).strftime("%b")
        all_months.append(mk)

    monthly_rev_values = [monthly_revenue.get((i+1, mk), 0) for i, mk in enumerate(all_months)]
    monthly_target_line = [PREMIUM_TARGET / 12] * len(all_months)

    # ── Conversion Funnel ──
    stage_order = ["New Lead", "Contacted", "Discovery Call", "Needs Analysis", "Plan Presentation", "Proposal Sent", "Negotiation", "Closed-Won"]
    funnel_counts = {}
    for s in stage_order:
        funnel_counts[s] = 0
    for p in prospects:
        s = p["stage"]
        if s in funnel_counts:
            funnel_counts[s] += 1
        # Count won deals as having passed through all prior stages
        if s == "Closed-Won":
            for prior in stage_order:
                funnel_counts[prior] += 1

    # Average days in each stage (for velocity)
    stage_days = {}
    for p in active:
        s = p["stage"]
        fc = p.get("first_contact", "")
        if fc and fc != "None" and s:
            try:
                start = datetime.strptime(fc.split(" ")[0], "%Y-%m-%d").date()
                days_in = (today - start).days
                if s not in stage_days:
                    stage_days[s] = []
                stage_days[s].append(days_in)
            except (ValueError, IndexError):
                pass

    avg_stage_days = {}
    for s, days_list in stage_days.items():
        avg_stage_days[s] = sum(days_list) / len(days_list)

    # Conversion rates between stages
    funnel_rates = []
    for i in range(len(stage_order) - 1):
        current = funnel_counts[stage_order[i]]
        next_s = funnel_counts[stage_order[i + 1]]
        rate = (next_s / current * 100) if current > 0 else 0
        funnel_rates.append(rate)

    # Source effectiveness
    source_wins = {}
    source_totals = {}
    for p in prospects:
        src = p["source"] or "Unknown"
        if p["stage"]:
            source_totals[src] = source_totals.get(src, 0) + 1
            if p["stage"] == "Closed-Won":
                source_wins[src] = source_wins.get(src, 0) + 1

    source_conversion = {}
    for src, total in source_totals.items():
        wins = source_wins.get(src, 0)
        source_conversion[src] = {"wins": wins, "total": total, "rate": (wins / total * 100) if total > 0 else 0}

    # ── Activity Scoreboard ──
    # Count activities this week and today
    week_start = today - timedelta(days=today.weekday())  # Monday
    activities_today = 0
    activities_week = 0
    calls_today = 0
    calls_week = 0
    emails_today = 0
    emails_week = 0
    meetings_today = 0
    meetings_week = 0

    for a in activities:
        ad = a["date"]
        if ad and ad != "None":
            try:
                activity_date = datetime.strptime(ad.split(" ")[0], "%Y-%m-%d").date()
                action = a["action"].lower()
                is_today = activity_date == today
                is_week = activity_date >= week_start

                if is_today:
                    activities_today += 1
                    if "call" in action or "phone" in action:
                        calls_today += 1
                    if "email" in action:
                        emails_today += 1
                    if "meeting" in action or "discovery" in action or "presentation" in action:
                        meetings_today += 1

                if is_week:
                    activities_week += 1
                    if "call" in action or "phone" in action:
                        calls_week += 1
                    if "email" in action:
                        emails_week += 1
                    if "meeting" in action or "discovery" in action or "presentation" in action:
                        meetings_week += 1
            except (ValueError, IndexError):
                pass

    # Include insurance book calls in call counts
    for b in book_entries:
        lc = b.get("last_called", "")
        if lc and lc != "None":
            try:
                call_date = datetime.strptime(lc.split(" ")[0], "%Y-%m-%d").date()
                if call_date == today:
                    calls_today += 1
                if call_date >= week_start:
                    calls_week += 1
            except (ValueError, IndexError):
                pass

    # Daily targets
    DAILY_CALLS_TARGET = 10
    DAILY_EMAILS_TARGET = 3
    WEEKLY_MEETINGS_TARGET = 5

    # Calculate streaks (consecutive days with at least 1 activity)
    activity_dates = set()
    for a in activities:
        ad = a["date"]
        if ad and ad != "None":
            try:
                activity_dates.add(datetime.strptime(ad.split(" ")[0], "%Y-%m-%d").date())
            except (ValueError, IndexError):
                pass
    for b in book_entries:
        lc = b.get("last_called", "")
        if lc and lc != "None":
            try:
                activity_dates.add(datetime.strptime(lc.split(" ")[0], "%Y-%m-%d").date())
            except (ValueError, IndexError):
                pass

    streak = 0
    check_date = today
    while check_date in activity_dates:
        streak += 1
        check_date -= timedelta(days=1)

    # Pre-compute deals aging rows
    aging_rows = ""
    for p in sorted(active, key=lambda x: x.get("first_contact", "9999")):
        fc = p.get("first_contact", "")
        if fc and fc != "None":
            try:
                days_open = (today - datetime.strptime(fc.split(" ")[0], "%Y-%m-%d").date()).days
                stale = '<span class="overdue">Stale</span>' if days_open > 30 else "OK"
            except (ValueError, IndexError):
                days_open = "?"
                stale = "?"
        else:
            days_open = "?"
            stale = "?"
        stage_bg = STAGE_COLORS.get(p["stage"], "#BDC3C7")
        aging_rows += f'<tr><td class="name-cell">{_esc(p["name"])}</td><td><span class="badge" style="background:{stage_bg}">{_esc(p["stage"])}</span></td><td>{days_open}</td><td>{stale}</td></tr>'

    # Pre-compute source effectiveness rows
    source_eff_rows = ""
    for src, data in sorted(source_conversion.items(), key=lambda x: -x[1]["rate"]):
        if data["total"] >= 1:
            rate_bg = "#27ae60" if data["rate"] >= 30 else "#f39c12" if data["rate"] >= 15 else "#e74c3c"
            source_eff_rows += f'<tr><td>{_esc(src)}</td><td>{data["total"]}</td><td>{data["wins"]}</td><td><span class="badge" style="background:{rate_bg}">{data["rate"]:.0f}%</span></td></tr>'

    # Stage counts for chart
    stage_counts = {}
    stage_revenue = {}
    stage_fyc = {}
    for p in prospects:
        s = p["stage"]
        if s:
            stage_counts[s] = stage_counts.get(s, 0) + 1
            try:
                stage_revenue[s] = stage_revenue.get(s, 0) + parse_money(p["revenue"])
            except ValueError:
                pass
            stage_fyc[s] = stage_fyc.get(s, 0) + calc_fyc(p["revenue"], p["product"])

    # Source counts
    source_counts = {}
    for p in prospects:
        s = p["source"]
        if s:
            source_counts[s] = source_counts.get(s, 0) + 1

    # Product counts
    product_counts = {}
    for p in prospects:
        pr = p["product"]
        if pr:
            product_counts[pr] = product_counts.get(pr, 0) + 1

    # Build prospect rows
    prospect_rows = ""
    for p in active:
        stage_bg = STAGE_COLORS.get(p["stage"], "#BDC3C7")
        stage_fg = STAGE_TEXT.get(p["stage"], "#fff")
        pri_bg = PRIORITY_COLORS.get(p["priority"], "#BDC3C7")

        is_overdue = p in overdue
        fu_class = "overdue" if is_overdue else ""
        fu_display = p["next_followup"].split(" ")[0] if p["next_followup"] and p["next_followup"] != "None" else ""

        p_json_escaped = _esc(json.dumps(p))
        prospect_rows += f"""<tr class="editable-row" data-prospect="{p_json_escaped}" onclick="openEdit(JSON.parse(this.dataset.prospect))" style="cursor:pointer">
            <td class="name-cell">{_esc(p["name"])}</td>
            <td><span class="badge" style="background:{pri_bg}">{_esc(p["priority"])}</span></td>
            <td><span class="badge" style="background:{stage_bg};color:{stage_fg}">{_esc(p["stage"])}</span></td>
            <td>{_esc(p["product"])}</td>
            <td class="money">{fmt_money_full(p["aum"])}</td>
            <td class="money">{fmt_money_full(p["revenue"])}</td>
            <td class="{fu_class}">{_esc(fu_display)}</td>
            <td class="notes">{_esc(p["notes"][:60])}{'...' if len(p["notes"]) > 60 else ''}</td>
        </tr>"""

    # Won deals rows
    won_rows = ""
    for p in won:
        won_rows += f"""<tr>
            <td class="name-cell">{_esc(p["name"])}</td>
            <td>{_esc(p["product"])}</td>
            <td class="money">{fmt_money_full(p["aum"])}</td>
            <td class="money">{fmt_money_full(p["revenue"])}</td>
            <td>{_esc(p["source"])}</td>
        </tr>"""

    # Activity rows (last 10)
    activity_rows = ""
    for a in activities[:10]:
        activity_rows += f"""<tr>
            <td>{_esc(a["date"].split(" ")[0])}</td>
            <td>{_esc(a["prospect"])}</td>
            <td>{_esc(a["action"])}</td>
            <td>{_esc(a["outcome"])}</td>
            <td>{_esc(a["next_step"])}</td>
        </tr>"""

    # Overdue rows
    overdue_rows = ""
    for p in overdue:
        fu = p["next_followup"].split(" ")[0] if p["next_followup"] else ""
        try:
            days_late = (today - datetime.strptime(fu, "%Y-%m-%d").date()).days
        except (ValueError, IndexError):
            days_late = "?"
        overdue_rows += f"""<tr>
            <td class="name-cell">{_esc(p["name"])}</td>
            <td>{_esc(fu)}</td>
            <td class="overdue">{days_late} days late</td>
            <td>{_esc(p["phone"])}</td>
        </tr>"""

    # Chart data as JSON-like strings for inline JS
    stage_labels = list(stage_counts.keys())
    stage_values = list(stage_counts.values())
    stage_chart_colors = [STAGE_COLORS.get(s, "#BDC3C7") for s in stage_labels]

    source_labels = list(source_counts.keys())
    source_values = list(source_counts.values())

    product_labels = list(product_counts.keys())
    product_values = list(product_counts.values())

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Calm Money — Pipeline Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
:root {{
    --bg: #0a0e17;
    --surface: #111827;
    --surface-2: #1a2332;
    --surface-3: #1f2937;
    --border: #2a3544;
    --text: #e2e8f0;
    --text-muted: #64748b;
    --text-dim: #475569;
    --accent: #10b981;
    --accent-glow: rgba(16, 185, 129, 0.15);
    --blue: #3b82f6;
    --blue-glow: rgba(59, 130, 246, 0.15);
    --gold: #f59e0b;
    --gold-glow: rgba(245, 158, 11, 0.15);
    --purple: #8b5cf6;
    --purple-glow: rgba(139, 92, 246, 0.15);
    --red: #ef4444;
    --red-glow: rgba(239, 68, 68, 0.15);
    --teal: #14b8a6;
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'DM Sans', -apple-system, sans-serif; background: var(--bg); color: var(--text); }}

.header {{
    background: var(--surface);
    padding: 20px 32px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    border-bottom: 1px solid var(--border);
}}
.header h1 {{ font-size: 20px; font-weight: 700; letter-spacing: -0.3px; color: var(--text); }}
.header h1 span {{ color: var(--accent); }}
.header .updated {{ font-size: 12px; color: var(--text-muted); font-family: 'JetBrains Mono', monospace; font-size: 11px; }}

.container {{ max-width: 1440px; margin: 0 auto; padding: 24px; }}

/* ── KPI Cards ── */
.kpi-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 12px;
    margin-bottom: 20px;
}}
.kpi-card {{
    background: var(--surface);
    border-radius: 10px;
    padding: 18px 20px;
    border: 1px solid var(--border);
    border-left: 3px solid var(--accent);
    transition: border-color 0.2s, box-shadow 0.2s;
}}
.kpi-card:hover {{ box-shadow: 0 0 20px var(--accent-glow); }}
.kpi-card.blue {{ border-left-color: var(--blue); }}
.kpi-card.blue:hover {{ box-shadow: 0 0 20px var(--blue-glow); }}
.kpi-card.green {{ border-left-color: var(--accent); }}
.kpi-card.green:hover {{ box-shadow: 0 0 20px var(--accent-glow); }}
.kpi-card.purple {{ border-left-color: var(--purple); }}
.kpi-card.purple:hover {{ box-shadow: 0 0 20px var(--purple-glow); }}
.kpi-card.red {{ border-left-color: var(--red); }}
.kpi-card.red:hover {{ box-shadow: 0 0 20px var(--red-glow); }}
.kpi-card.gold {{ border-left-color: var(--gold); }}
.kpi-card.gold:hover {{ box-shadow: 0 0 20px var(--gold-glow); }}
.kpi-label {{ font-size: 10px; text-transform: uppercase; color: var(--text-muted); font-weight: 600; letter-spacing: 1px; }}
.kpi-value {{ font-size: 28px; font-weight: 700; margin-top: 4px; color: var(--text); font-family: 'JetBrains Mono', monospace; letter-spacing: -0.5px; }}

/* ── Sections ── */
.section {{
    background: var(--surface);
    border-radius: 10px;
    padding: 24px;
    margin-bottom: 20px;
    border: 1px solid var(--border);
}}
.section h2 {{
    font-size: 13px;
    font-weight: 600;
    color: var(--text);
    margin-bottom: 16px;
    padding-bottom: 10px;
    border-bottom: 1px solid var(--border);
    text-transform: uppercase;
    letter-spacing: 0.5px;
}}
.section h2 .count {{ color: var(--text-muted); font-weight: 400; }}

/* ── Charts ── */
.chart-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 16px;
    margin-bottom: 20px;
}}
.chart-card {{
    background: var(--surface);
    border-radius: 10px;
    padding: 20px;
    border: 1px solid var(--border);
}}
.chart-card h3 {{ font-size: 12px; font-weight: 600; color: var(--text-muted); margin-bottom: 12px; text-transform: uppercase; letter-spacing: 0.5px; }}

/* ── Tables ── */
table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
th {{ text-align: left; padding: 10px 12px; background: var(--surface-2); color: var(--text-muted); font-weight: 600; font-size: 10px; text-transform: uppercase; letter-spacing: 0.8px; border-bottom: 1px solid var(--border); }}
td {{ padding: 10px 12px; border-bottom: 1px solid var(--border); color: var(--text); }}
tr:hover {{ background: var(--surface-2); }}

.badge {{
    display: inline-block;
    padding: 3px 10px;
    border-radius: 4px;
    font-size: 10px;
    font-weight: 600;
    color: white;
    text-transform: uppercase;
    letter-spacing: 0.3px;
}}
.name-cell {{ font-weight: 600; color: var(--text); }}
.money {{ font-family: 'JetBrains Mono', monospace; text-align: right; font-size: 12px; }}
.notes {{ color: var(--text-muted); font-size: 12px; max-width: 200px; }}
.overdue {{ color: var(--red); font-weight: 600; }}

.two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}

.empty-state {{ text-align: center; padding: 40px; color: var(--text-dim); }}
.empty-state p {{ margin-top: 8px; font-size: 13px; }}

.refresh-note {{ text-align: center; color: var(--text-dim); font-size: 11px; margin-top: 16px; padding: 12px; }}

.editable-row:hover {{ background: var(--surface-3) !important; cursor: pointer; }}

/* ── Tabs ── */
.tab-nav {{
    display: flex;
    gap: 0;
    background: var(--surface);
    border-radius: 10px 10px 0 0;
    margin-bottom: 0;
    border: 1px solid var(--border);
    border-bottom: none;
    overflow: hidden;
}}
.tab-btn {{
    flex: 1;
    padding: 14px 20px;
    border: none;
    background: transparent;
    font-family: 'DM Sans', sans-serif;
    font-size: 12px;
    font-weight: 600;
    color: var(--text-muted);
    cursor: pointer;
    border-bottom: 2px solid transparent;
    transition: all 0.2s;
    text-transform: uppercase;
    letter-spacing: 0.8px;
}}
.tab-btn:hover {{ background: var(--surface-2); color: var(--text); }}
.tab-btn.active {{ color: var(--accent); border-bottom-color: var(--accent); background: var(--surface-2); }}
.tab-content {{ display: none; }}
.tab-content.active {{ display: block; }}

/* ── Progress bars ── */
.progress-bar-container {{
    background: var(--surface-3);
    border-radius: 6px;
    height: 22px;
    overflow: hidden;
    position: relative;
    margin: 8px 0;
}}
.progress-bar-fill {{
    height: 100%;
    border-radius: 6px;
    transition: width 0.8s cubic-bezier(0.4, 0, 0.2, 1);
    display: flex;
    align-items: center;
    padding-left: 8px;
    font-size: 10px;
    font-weight: 600;
    color: white;
    min-width: 36px;
    font-family: 'JetBrains Mono', monospace;
}}
.progress-bar-fill.green {{ background: linear-gradient(90deg, #059669, #10b981); }}
.progress-bar-fill.red {{ background: linear-gradient(90deg, #dc2626, #ef4444); }}
.progress-bar-fill.blue {{ background: linear-gradient(90deg, #2563eb, #3b82f6); }}
.progress-bar-fill.teal {{ background: linear-gradient(90deg, #0d9488, #14b8a6); }}

.pace-indicator {{
    display: inline-block;
    padding: 3px 10px;
    border-radius: 4px;
    font-size: 10px;
    font-weight: 600;
    font-family: 'JetBrains Mono', monospace;
}}
.pace-ahead {{ background: rgba(16,185,129,0.15); color: var(--accent); }}
.pace-behind {{ background: rgba(239,68,68,0.15); color: var(--red); }}

.target-card {{
    background: var(--surface);
    border-radius: 10px;
    padding: 24px;
    border: 1px solid var(--border);
    margin-bottom: 16px;
}}
.target-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 12px;
}}
.target-header h3 {{
    font-size: 14px;
    font-weight: 700;
    color: var(--text);
}}
.target-meta {{
    display: flex;
    justify-content: space-between;
    font-size: 11px;
    color: var(--text-muted);
    margin-top: 4px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px;
}}

/* ── Funnel ── */
.funnel-stage {{
    display: flex;
    align-items: center;
    margin-bottom: 6px;
    gap: 12px;
}}
.funnel-label {{
    width: 140px;
    font-size: 11px;
    font-weight: 600;
    color: var(--text);
    text-align: right;
}}
.funnel-bar-wrap {{
    flex: 1;
    display: flex;
    align-items: center;
    gap: 8px;
}}
.funnel-bar {{
    height: 26px;
    border-radius: 4px;
    display: flex;
    align-items: center;
    padding-left: 10px;
    font-size: 11px;
    font-weight: 600;
    color: white;
    min-width: 28px;
    transition: width 0.6s cubic-bezier(0.4, 0, 0.2, 1);
    font-family: 'JetBrains Mono', monospace;
}}
.funnel-rate {{ font-size: 10px; color: var(--text-muted); white-space: nowrap; font-family: 'JetBrains Mono', monospace; }}
.funnel-velocity {{ font-size: 10px; color: var(--text-dim); min-width: 60px; font-family: 'JetBrains Mono', monospace; }}

/* ── Scoreboard ── */
.score-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 12px;
    margin-bottom: 20px;
}}
.score-card {{
    background: var(--surface);
    border-radius: 10px;
    padding: 20px;
    text-align: center;
    border: 1px solid var(--border);
    border-top: 3px solid var(--accent);
}}
.score-card.fire {{ border-top-color: var(--red); }}
.score-card h4 {{
    font-size: 10px;
    text-transform: uppercase;
    color: var(--text-muted);
    letter-spacing: 1px;
    margin-bottom: 8px;
}}
.score-big {{
    font-size: 32px;
    font-weight: 700;
    color: var(--text);
    font-family: 'JetBrains Mono', monospace;
}}
.score-target {{
    font-size: 11px;
    color: var(--text-muted);
    margin-top: 4px;
}}
.streak-badge {{
    display: inline-block;
    background: linear-gradient(135deg, var(--red), var(--gold));
    color: white;
    padding: 6px 16px;
    border-radius: 4px;
    font-size: 13px;
    font-weight: 700;
    font-family: 'JetBrains Mono', monospace;
}}

.btn {{ display: inline-block; padding: 8px 20px; border-radius: 6px; font-size: 12px; font-weight: 600; border: none; cursor: pointer; font-family: 'DM Sans', sans-serif; letter-spacing: 0.3px; }}
.btn-primary {{ background: var(--accent); color: #000; }}
.btn-primary:hover {{ background: #059669; }}
.btn-danger {{ background: var(--red); color: white; }}
.btn-danger:hover {{ background: #dc2626; }}
.btn-secondary {{ background: var(--surface-3); color: var(--text); border: 1px solid var(--border); }}

.modal-overlay {{ display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.7); backdrop-filter: blur(4px); z-index: 1000; justify-content: center; align-items: center; }}
.modal-overlay.active {{ display: flex; }}
.modal {{ background: var(--surface); border: 1px solid var(--border); border-radius: 12px; padding: 32px; width: 500px; max-width: 90vw; max-height: 90vh; overflow-y: auto; box-shadow: 0 24px 80px rgba(0,0,0,0.5); }}
.modal h2 {{ font-size: 16px; margin-bottom: 20px; color: var(--text); }}
.modal label {{ display: block; font-size: 10px; font-weight: 600; text-transform: uppercase; color: var(--text-muted); margin-bottom: 4px; margin-top: 12px; letter-spacing: 0.8px; }}
.modal input, .modal select, .modal textarea {{ width: 100%; padding: 8px 12px; border: 1px solid var(--border); border-radius: 6px; font-size: 13px; font-family: 'DM Sans', sans-serif; background: var(--surface-2); color: var(--text); }}
.modal textarea {{ resize: vertical; min-height: 60px; }}
.modal select {{ background: var(--surface-2); color: var(--text); }}
.modal .form-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
.modal .actions {{ display: flex; gap: 8px; margin-top: 24px; justify-content: flex-end; }}
.modal .actions .left {{ margin-right: auto; }}

.add-btn {{ margin-bottom: 16px; float: right; }}

@media (max-width: 900px) {{
    .chart-grid {{ grid-template-columns: 1fr; }}
    .two-col {{ grid-template-columns: 1fr; }}
    .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }}
}}
</style>
</head>
<body>

<div class="header">
    <div>
        <h1>CALM <span>MONEY</span> — Pipeline</h1>
    </div>
    <div class="updated">Updated: {now.strftime('%B %d, %Y at %I:%M %p')}<br>Refresh page for latest data</div>
</div>

<div class="container">

    <div class="kpi-grid" style="grid-template-columns: repeat(4, 1fr)">
        <div class="kpi-card blue">
            <div class="kpi-label">Total AUM</div>
            <div class="kpi-value">{fmt_money(forecast_aum)}</div>
        </div>
        <div class="kpi-card green">
            <div class="kpi-label">Premium YTD</div>
            <div class="kpi-value">{fmt_money(forecast_revenue)}</div>
        </div>
        <div class="kpi-card gold">
            <div class="kpi-label">FYC YTD</div>
            <div class="kpi-value">{fmt_money(won_fyc)}</div>
        </div>
        <div class="kpi-card purple">
            <div class="kpi-label">Pipeline AUM</div>
            <div class="kpi-value">{fmt_money(total_pipeline)}</div>
        </div>
    </div>
    <div class="kpi-grid" style="grid-template-columns: repeat(4, 1fr); margin-top: 12px">
        <div class="kpi-card">
            <div class="kpi-label">Active Deals</div>
            <div class="kpi-value">{len(active)}</div>
        </div>
        <div class="kpi-card red">
            <div class="kpi-label">Hot Leads</div>
            <div class="kpi-value">{hot_count}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">Win Rate</div>
            <div class="kpi-value">{win_rate:.0f}%</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">Overdue</div>
            <div class="kpi-value">{len(overdue)}</div>
        </div>
    </div>

    <div class="tab-nav">
        <button class="tab-btn active" onclick="showTab('pipeline')">Pipeline</button>
        <button class="tab-btn" onclick="showTab('forecast')">Revenue Forecast</button>
        <button class="tab-btn" onclick="showTab('funnel')">Conversion Funnel</button>
        <button class="tab-btn" onclick="showTab('scoreboard')">Activity Score</button>
    </div>

    <!-- ═══ TAB 1: PIPELINE (existing) ═══ -->
    <div class="tab-content active" id="tab-pipeline">

    <div class="chart-grid" style="margin-top:24px">
        <div class="chart-card">
            <h3>By Stage</h3>
            <canvas id="stageChart"></canvas>
        </div>
        <div class="chart-card">
            <h3>By Source</h3>
            <canvas id="sourceChart"></canvas>
        </div>
        <div class="chart-card">
            <h3>By Product</h3>
            <canvas id="productChart"></canvas>
        </div>
    </div>

    {'<div class="section"><h2>Overdue Follow-Ups <span class="count">(' + str(len(overdue)) + ')</span></h2><table><tr><th>Prospect</th><th>Was Due</th><th>Status</th><th>Phone</th></tr>' + overdue_rows + '</table></div>' if overdue else ''}

    <div class="section">
        <h2 style="display:flex;justify-content:space-between;align-items:center">Active Pipeline <span class="count">({len(active)} deals)</span> <button class="btn btn-primary" onclick="openAdd()">+ Add Prospect</button></h2>
        {'<table><tr><th>Prospect</th><th>Priority</th><th>Stage</th><th>Product</th><th>AUM/Premium</th><th>Revenue</th><th>Follow-Up</th><th>Notes</th></tr>' + prospect_rows + '</table>' if active else '<div class="empty-state"><p>No active deals yet. Text your Telegram bot to add prospects.</p></div>'}
    </div>

    <div class="two-col">
        <div class="section">
            <h2>Closed-Won <span class="count">({len(won)})</span></h2>
            {'<table><tr><th>Client</th><th>Product</th><th>AUM</th><th>Revenue</th><th>Source</th></tr>' + won_rows + '</table>' if won else '<div class="empty-state"><p>No wins yet. Keep grinding!</p></div>'}
        </div>
        <div class="section">
            <h2>Recent Activity <span class="count">(last 10)</span></h2>
            {'<table><tr><th>Date</th><th>Prospect</th><th>Action</th><th>Outcome</th><th>Next</th></tr>' + activity_rows + '</table>' if activities else '<div class="empty-state"><p>No activity logged yet.</p></div>'}
        </div>
    </div>

    <div class="two-col">
        <div class="section">
            <h2>Upcoming Meetings <span class="count">({len([m for m in meetings if m['status'] != 'Cancelled'])})</span></h2>
            {'<table><tr><th>Date</th><th>Time</th><th>Prospect</th><th>Type</th><th>Status</th><th>Prep</th></tr>' + ''.join(f'<tr><td>{_esc(m["date"])}</td><td>{_esc(m["time"])}</td><td class="name-cell">{_esc(m["prospect"])}</td><td>{_esc(m["type"])}</td><td><span class="badge" style="background:{"#27ae60" if m["status"]=="Completed" else "#e74c3c" if m["status"]=="Cancelled" else "#3498db"}">{_esc(m["status"])}</span></td><td class="notes">{_esc(m["prep_notes"][:50])}{"..." if len(m["prep_notes"])>50 else ""}</td></tr>' for m in meetings if m['status'] != 'Cancelled') + '</table>' if meetings else '<div class="empty-state"><p>No meetings scheduled. Text the bot to add one.</p></div>'}
        </div>
        <div class="section">
            <h2>Insurance Book <span class="count">({len(book_entries)} contacts)</span></h2>
            {'<div style="display:flex;gap:24px;margin-bottom:16px"><div class="kpi-card" style="flex:1;padding:12px 16px"><div class="kpi-label">Called</div><div class="kpi-value" style="font-size:24px">' + str(len([b for b in book_entries if b["status"].lower() not in ("not called","")])) + '</div></div><div class="kpi-card green" style="flex:1;padding:12px 16px"><div class="kpi-label">Booked</div><div class="kpi-value" style="font-size:24px">' + str(len([b for b in book_entries if b["status"].lower()=="booked meeting"])) + '</div></div><div class="kpi-card blue" style="flex:1;padding:12px 16px"><div class="kpi-label">Remaining</div><div class="kpi-value" style="font-size:24px">' + str(len([b for b in book_entries if b["status"].lower() in ("not called","")])) + '</div></div></div><table><tr><th>Name</th><th>Phone</th><th>Status</th><th>Last Called</th><th>Notes</th></tr>' + ''.join(f'<tr><td class="name-cell">{_esc(b["name"])}</td><td>{_esc(b["phone"])}</td><td><span class="badge" style="background:{"#27ae60" if b["status"].lower()=="booked meeting" else "#e74c3c" if b["status"].lower()=="not interested" else "#f39c12" if b["status"].lower() in ("callback","no answer") else "#3498db"}">{_esc(b["status"])}</span></td><td>{_esc(b["last_called"].split(" ")[0] if b["last_called"] and b["last_called"]!="None" else "")}</td><td class="notes">{_esc(b["notes"][:40])}{"..." if len(b["notes"])>40 else ""}</td></tr>' for b in book_entries[:20]) + '</table>' if book_entries else '<div class="empty-state"><p>No insurance book uploaded. Send a CSV via Telegram.</p></div>'}
        </div>
    </div>

    <div class="refresh-note">Click any prospect row to edit. Changes save to your pipeline instantly.</div>

    </div><!-- end tab-pipeline -->

    <!-- ═══ TAB 2: REVENUE FORECAST ═══ -->
    <div class="tab-content" id="tab-forecast" style="margin-top:24px">

        <div class="kpi-grid" style="grid-template-columns: repeat(3, 1fr)">
            <div class="kpi-card green">
                <div class="kpi-label">Total Premium YTD</div>
                <div class="kpi-value">{fmt_money(forecast_revenue)}</div>
            </div>
            <div class="kpi-card blue">
                <div class="kpi-label">Total AUM</div>
                <div class="kpi-value">{fmt_money(forecast_aum)}</div>
            </div>
            <div class="kpi-card gold">
                <div class="kpi-label">FYC (Won)</div>
                <div class="kpi-value">{fmt_money(won_fyc)}</div>
            </div>
        </div>
        <div class="kpi-grid" style="grid-template-columns: repeat(3, 1fr); margin-top:12px">
            <div class="kpi-card purple">
                <div class="kpi-label">Projected Premium</div>
                <div class="kpi-value">{fmt_money(projected_revenue)}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Projected AUM</div>
                <div class="kpi-value">{fmt_money(projected_aum)}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Projected FYC</div>
                <div class="kpi-value">{fmt_money(projected_fyc)}</div>
            </div>
        </div>

        <div class="two-col">
            <div class="target-card">
                <div class="target-header">
                    <h3>Premium Target: {fmt_money(PREMIUM_TARGET)}</h3>
                    <span class="pace-indicator {'pace-ahead' if premium_on_pace else 'pace-behind'}">{'Ahead of pace' if premium_on_pace else 'Behind pace'}</span>
                </div>
                <div class="progress-bar-container">
                    <div class="progress-bar-fill {'green' if premium_on_pace else 'red'}" style="width:{min(premium_pct, 100):.0f}%">{premium_pct:.0f}%</div>
                </div>
                <div class="target-meta">
                    <span>Actual: {fmt_money(forecast_revenue)} (baseline {fmt_money(BASELINE_PREMIUM)} + pipeline {fmt_money(won_revenue)})</span>
                    <span>Gap: {fmt_money(max(0, PREMIUM_TARGET - forecast_revenue))}</span>
                </div>
                <div class="target-meta" style="margin-top:8px">
                    <span>Pipeline weighted: +{fmt_money(weighted_revenue)}</span>
                    <span>Need {fmt_money(max(0, (PREMIUM_TARGET - forecast_revenue - weighted_revenue) / max(1, days_remaining) * 30))}/mo to close gap</span>
                </div>
            </div>
            <div class="target-card">
                <div class="target-header">
                    <h3>AUM Target: {fmt_money(AUM_TARGET)}</h3>
                    <span class="pace-indicator {'pace-ahead' if aum_on_pace else 'pace-behind'}">{'Ahead of pace' if aum_on_pace else 'Behind pace'}</span>
                </div>
                <div class="progress-bar-container">
                    <div class="progress-bar-fill {'green' if aum_on_pace else 'red'}" style="width:{min(aum_pct, 100):.0f}%">{aum_pct:.0f}%</div>
                </div>
                <div class="target-meta">
                    <span>Actual: {fmt_money(forecast_aum)} (baseline {fmt_money(BASELINE_AUM)} + pipeline {fmt_money(won_aum_pipeline)})</span>
                    <span>Gap: {fmt_money(max(0, AUM_TARGET - forecast_aum))}</span>
                </div>
                <div class="target-meta" style="margin-top:8px">
                    <span>Pipeline weighted: +{fmt_money(weighted_aum)}</span>
                    <span>Need {fmt_money(max(0, (AUM_TARGET - forecast_aum - weighted_aum) / max(1, days_remaining) * 30))}/mo to close gap</span>
                </div>
            </div>
        </div>

        <div class="target-card">
            <h3 style="margin-bottom:16px">Monthly Revenue vs Target</h3>
            <canvas id="monthlyChart" height="80"></canvas>
        </div>

        <div class="target-card">
            <div class="target-header">
                <h3>Year Progress</h3>
                <span style="font-size:12px;color:var(--text-muted)">{days_elapsed} of {days_total} days ({pct_year:.0f}%)</span>
            </div>
            <div class="progress-bar-container">
                <div class="progress-bar-fill teal" style="width:{pct_year:.0f}%">{pct_year:.0f}%</div>
            </div>
        </div>

        <div class="section">
            <h2>Pipeline Weighted Forecast</h2>
            <table>
                <tr><th>Stage</th><th>Deals</th><th>Prob</th><th>Premium</th><th>Wtd Premium</th><th>AUM</th><th>Wtd AUM</th><th>FYC</th><th>Wtd FYC</th></tr>
                {''.join(f'<tr><td><span class="badge" style="background:{STAGE_COLORS.get(s, "#BDC3C7")}">{_esc(s)}</span></td><td>{stage_counts.get(s, 0)}</td><td>{int(stage_probability.get(s, 0.1)*100)}%</td><td class="money">{fmt_money(stage_revenue.get(s, 0))}</td><td class="money">{fmt_money(stage_revenue.get(s, 0) * stage_probability.get(s, 0.1))}</td><td class="money">{fmt_money(sum(parse_money(p["aum"]) for p in active if p["stage"]==s))}</td><td class="money">{fmt_money(sum(parse_money(p["aum"]) for p in active if p["stage"]==s) * stage_probability.get(s, 0.1))}</td><td class="money">{fmt_money(stage_fyc.get(s, 0))}</td><td class="money">{fmt_money(stage_fyc.get(s, 0) * stage_probability.get(s, 0.1))}</td></tr>' for s in stage_order[:-1] if stage_counts.get(s, 0) > 0)}
                <tr style="font-weight:700;border-top:2px solid var(--border)"><td>Total Weighted</td><td></td><td></td><td></td><td class="money">{fmt_money(weighted_revenue)}</td><td></td><td class="money">{fmt_money(weighted_aum)}</td><td></td><td class="money">{fmt_money(weighted_fyc)}</td></tr>
            </table>
        </div>

    </div><!-- end tab-forecast -->

    <!-- ═══ TAB 3: CONVERSION FUNNEL ═══ -->
    <div class="tab-content" id="tab-funnel" style="margin-top:24px">

        <div class="section">
            <h2>Sales Funnel</h2>
            <div style="max-width:700px;margin:0 auto;padding:20px 0">
                {''.join(f'<div class="funnel-stage"><div class="funnel-label">{_esc(stage_order[i])}</div><div class="funnel-bar-wrap"><div class="funnel-bar" style="width:{max(8, funnel_counts[stage_order[i]] / max(1, funnel_counts[stage_order[0]]) * 100):.0f}%;background:{STAGE_COLORS.get(stage_order[i], "#BDC3C7")}">{funnel_counts[stage_order[i]]}</div><div class="funnel-rate">{f"{funnel_rates[i]:.0f}% pass" if i < len(funnel_rates) else ""}</div><div class="funnel-velocity">{f"~{avg_stage_days.get(stage_order[i], 0):.0f}d avg" if stage_order[i] in avg_stage_days else ""}</div></div></div>' for i in range(len(stage_order)))}
            </div>
        </div>

        <div class="two-col">
            <div class="section">
                <h2>Source Effectiveness</h2>
                {'<table><tr><th>Source</th><th>Total Leads</th><th>Wins</th><th>Conversion</th></tr>' + source_eff_rows + '</table>' if source_conversion else '<div class="empty-state"><p>No source data yet.</p></div>'}
            </div>
            <div class="section">
                <h2>Deals Aging (Active Pipeline)</h2>
                {'<table><tr><th>Prospect</th><th>Stage</th><th>Days Open</th><th>Status</th></tr>' + aging_rows + '</table>' if active else '<div class="empty-state"><p>No active deals.</p></div>'}
            </div>
        </div>

        <div class="section">
            <h2>Stage Velocity</h2>
            <canvas id="velocityChart" height="60"></canvas>
        </div>

    </div><!-- end tab-funnel -->

    <!-- ═══ TAB 4: ACTIVITY SCOREBOARD ═══ -->
    <div class="tab-content" id="tab-scoreboard" style="margin-top:24px">

        <div style="text-align:center;margin-bottom:24px">
            {f'<div class="streak-badge">🔥 {streak} Day Streak</div>' if streak > 0 else '<div style="color:var(--text-muted);font-size:14px">No streak yet — make a call to start one!</div>'}
        </div>

        <div class="score-grid">
            <div class="score-card {'fire' if calls_today >= DAILY_CALLS_TARGET else ''}">
                <h4>Calls Today</h4>
                <div class="score-big">{calls_today}</div>
                <div class="score-target">Target: {DAILY_CALLS_TARGET}</div>
                <div class="progress-bar-container" style="margin-top:8px">
                    <div class="progress-bar-fill {'green' if calls_today >= DAILY_CALLS_TARGET else 'blue'}" style="width:{min(calls_today / DAILY_CALLS_TARGET * 100, 100):.0f}%"></div>
                </div>
            </div>
            <div class="score-card {'fire' if emails_today >= DAILY_EMAILS_TARGET else ''}">
                <h4>Emails Today</h4>
                <div class="score-big">{emails_today}</div>
                <div class="score-target">Target: {DAILY_EMAILS_TARGET}</div>
                <div class="progress-bar-container" style="margin-top:8px">
                    <div class="progress-bar-fill {'green' if emails_today >= DAILY_EMAILS_TARGET else 'blue'}" style="width:{min(emails_today / DAILY_EMAILS_TARGET * 100, 100):.0f}%"></div>
                </div>
            </div>
            <div class="score-card {'fire' if meetings_week >= WEEKLY_MEETINGS_TARGET else ''}">
                <h4>Meetings This Week</h4>
                <div class="score-big">{meetings_week}</div>
                <div class="score-target">Target: {WEEKLY_MEETINGS_TARGET}</div>
                <div class="progress-bar-container" style="margin-top:8px">
                    <div class="progress-bar-fill {'green' if meetings_week >= WEEKLY_MEETINGS_TARGET else 'blue'}" style="width:{min(meetings_week / WEEKLY_MEETINGS_TARGET * 100, 100):.0f}%"></div>
                </div>
            </div>
            <div class="score-card">
                <h4>Total Activities Today</h4>
                <div class="score-big">{activities_today}</div>
                <div class="score-target">Week total: {activities_week}</div>
            </div>
        </div>

        <div class="two-col">
            <div class="section">
                <h2>This Week's Numbers</h2>
                <table>
                    <tr><th>Metric</th><th>Today</th><th>This Week</th><th>Target</th></tr>
                    <tr><td>Calls</td><td>{calls_today}</td><td>{calls_week}</td><td>{DAILY_CALLS_TARGET}/day</td></tr>
                    <tr><td>Emails</td><td>{emails_today}</td><td>{emails_week}</td><td>{DAILY_EMAILS_TARGET}/day</td></tr>
                    <tr><td>Meetings</td><td>{meetings_today}</td><td>{meetings_week}</td><td>{WEEKLY_MEETINGS_TARGET}/week</td></tr>
                    <tr style="font-weight:700;border-top:2px solid var(--border)"><td>Total Activities</td><td>{activities_today}</td><td>{activities_week}</td><td></td></tr>
                </table>
            </div>
            <div class="section">
                <h2>Insurance Book Progress</h2>
                {'<div style="text-align:center;padding:20px"><div class="score-big" style="font-size:48px">' + str(len([b for b in book_entries if b["status"].lower() not in ("not called","")])) + '<span style="font-size:20px;color:var(--text-muted)">/' + str(len(book_entries)) + '</span></div><div style="color:var(--text-muted);margin-top:4px">Contacts Called</div><div class="progress-bar-container" style="margin-top:12px"><div class="progress-bar-fill teal" style="width:' + str(min(len([b for b in book_entries if b["status"].lower() not in ("not called","")]) / max(1, len(book_entries)) * 100, 100)) + '%">' + str(int(len([b for b in book_entries if b["status"].lower() not in ("not called","")]) / max(1, len(book_entries)) * 100)) + '%</div></div><div class="target-meta" style="margin-top:8px"><span>Booked: ' + str(len([b for b in book_entries if b["status"].lower()=="booked meeting"])) + '</span><span>Not Interested: ' + str(len([b for b in book_entries if b["status"].lower()=="not interested"])) + '</span><span>Callbacks: ' + str(len([b for b in book_entries if b["status"].lower()=="callback"])) + '</span></div></div>' if book_entries else '<div class="empty-state"><p>Upload an insurance book CSV to track progress.</p></div>'}
            </div>
        </div>

    </div><!-- end tab-scoreboard -->

</div>

<!-- Edit Modal -->
<div class="modal-overlay" id="editModal">
<div class="modal">
    <h2 id="modalTitle">Edit Prospect</h2>
    <input type="hidden" id="origName">
    <div class="form-row">
        <div><label>Name</label><input id="fName" type="text"></div>
        <div><label>Phone</label><input id="fPhone" type="text"></div>
    </div>
    <div class="form-row">
        <div><label>Email</label><input id="fEmail" type="text"></div>
        <div><label>Source</label>
            <select id="fSource">
                <option value="">—</option>
                <option>Referral</option><option>Website</option><option>Social Media</option>
                <option>Seminar</option><option>Cold Outreach</option><option>LinkedIn</option>
                <option>Podcast</option><option>Networking</option><option>Centre of Influence</option><option>Other</option>
            </select>
        </div>
    </div>
    <div class="form-row">
        <div><label>Priority</label>
            <select id="fPriority">
                <option value="">—</option>
                <option>Hot</option><option>Warm</option><option>Cold</option>
            </select>
        </div>
        <div><label>Stage</label>
            <select id="fStage">
                <option value="">—</option>
                <option>New Lead</option><option>Contacted</option><option>Discovery Call</option>
                <option>Needs Analysis</option><option>Plan Presentation</option><option>Proposal Sent</option>
                <option>Negotiation</option><option>Closed-Won</option><option>Closed-Lost</option><option>Nurture</option>
            </select>
        </div>
    </div>
    <div class="form-row">
        <div><label>Product</label>
            <select id="fProduct">
                <option value="">—</option>
                <option>Life Insurance</option><option>Wealth Management</option><option>Life Insurance + Wealth</option>
                <option>Disability Insurance</option><option>Critical Illness</option><option>Group Benefits</option>
                <option>Estate Planning</option><option>Other</option>
            </select>
        </div>
        <div><label>Next Follow-Up</label><input id="fFollowup" type="date"></div>
    </div>
    <div class="form-row">
        <div><label>AUM / Premium</label><input id="fAum" type="text" placeholder="e.g. 500000"></div>
        <div><label>Revenue</label><input id="fRevenue" type="text" placeholder="e.g. 5000"></div>
    </div>
    <label>Notes</label>
    <textarea id="fNotes"></textarea>
    <div class="actions">
        <button class="btn btn-danger left" id="deleteBtn" onclick="deleteProspect()">Delete</button>
        <button class="btn btn-secondary" onclick="closeModal()">Cancel</button>
        <button class="btn btn-primary" onclick="saveProspect()">Save</button>
    </div>
</div>
</div>

<script>
// Tab switching
function showTab(name) {{
    document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.getElementById('tab-' + name).classList.add('active');
    event.target.classList.add('active');
    // Initialize charts when their tab is shown
    if (name === 'forecast' && !window._forecastInit) initForecastCharts();
    if (name === 'funnel' && !window._funnelInit) initFunnelCharts();
}}

// Dark mode chart defaults
Chart.defaults.color = '#94a3b8';
Chart.defaults.borderColor = '#2a3544';
const chartColors = ['#10b981','#3b82f6','#8b5cf6','#f59e0b','#ef4444','#14b8a6','#ec4899','#06b6d4','#64748b','#f97316'];

new Chart(document.getElementById('stageChart'), {{
    type: 'doughnut',
    data: {{
        labels: {stage_labels},
        datasets: [{{ data: {stage_values}, backgroundColor: {stage_chart_colors} }}]
    }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 8, font: {{ size: 11 }} }} }} }} }}
}});

new Chart(document.getElementById('sourceChart'), {{
    type: 'doughnut',
    data: {{
        labels: {source_labels},
        datasets: [{{ data: {source_values}, backgroundColor: chartColors }}]
    }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 8, font: {{ size: 11 }} }} }} }} }}
}});

new Chart(document.getElementById('productChart'), {{
    type: 'doughnut',
    data: {{
        labels: {product_labels},
        datasets: [{{ data: {product_values}, backgroundColor: chartColors }}]
    }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 8, font: {{ size: 11 }} }} }} }} }}
}});

// Modal logic
let isAdding = false;

function openEdit(p) {{
    isAdding = false;
    document.getElementById('modalTitle').textContent = 'Edit: ' + p.name;
    document.getElementById('origName').value = p.name;
    document.getElementById('fName').value = p.name;
    document.getElementById('fPhone').value = p.phone || '';
    document.getElementById('fEmail').value = p.email || '';
    document.getElementById('fSource').value = p.source || '';
    document.getElementById('fPriority').value = p.priority || '';
    document.getElementById('fStage').value = p.stage || '';
    document.getElementById('fProduct').value = p.product || '';
    document.getElementById('fAum').value = p.aum || '';
    document.getElementById('fRevenue').value = p.revenue || '';
    document.getElementById('fNotes').value = p.notes || '';
    let fu = p.next_followup || '';
    if (fu && fu !== 'None') {{
        fu = fu.split(' ')[0];
        if (/^\\d{{4}}-\\d{{2}}-\\d{{2}}$/.test(fu)) document.getElementById('fFollowup').value = fu;
        else document.getElementById('fFollowup').value = '';
    }} else document.getElementById('fFollowup').value = '';
    document.getElementById('deleteBtn').style.display = 'inline-block';
    document.getElementById('editModal').classList.add('active');
}}

function openAdd() {{
    isAdding = true;
    document.getElementById('modalTitle').textContent = 'Add Prospect';
    document.getElementById('origName').value = '';
    ['fName','fPhone','fEmail','fNotes','fAum','fRevenue','fFollowup'].forEach(id => document.getElementById(id).value = '');
    ['fSource','fPriority','fProduct'].forEach(id => document.getElementById(id).value = '');
    document.getElementById('fStage').value = 'New Lead';
    document.getElementById('deleteBtn').style.display = 'none';
    document.getElementById('editModal').classList.add('active');
}}

function closeModal() {{
    document.getElementById('editModal').classList.remove('active');
}}

function getFormData() {{
    return {{
        name: document.getElementById('fName').value.trim(),
        phone: document.getElementById('fPhone').value.trim(),
        email: document.getElementById('fEmail').value.trim(),
        source: document.getElementById('fSource').value,
        priority: document.getElementById('fPriority').value,
        stage: document.getElementById('fStage').value,
        product: document.getElementById('fProduct').value,
        aum: document.getElementById('fAum').value.trim(),
        revenue: document.getElementById('fRevenue').value.trim(),
        next_followup: document.getElementById('fFollowup').value,
        notes: document.getElementById('fNotes').value.trim(),
    }};
}}

async function saveProspect() {{
    const data = getFormData();
    if (!data.name) {{ alert('Name is required'); return; }}
    try {{
        let res;
        if (isAdding) {{
            res = await fetch('/api/prospect', {{ method: 'POST', headers: {{'Content-Type': 'application/json'}}, body: JSON.stringify(data) }});
        }} else {{
            const origName = document.getElementById('origName').value;
            res = await fetch('/api/prospect/' + encodeURIComponent(origName), {{ method: 'PUT', headers: {{'Content-Type': 'application/json'}}, body: JSON.stringify(data) }});
        }}
        const result = await res.json();
        if (result.ok) {{ closeModal(); location.reload(); }}
        else alert(result.error || 'Error saving');
    }} catch(e) {{ alert('Error: ' + e.message); }}
}}

async function deleteProspect() {{
    const name = document.getElementById('origName').value;
    if (!confirm('Delete ' + name + '?')) return;
    try {{
        const res = await fetch('/api/prospect/' + encodeURIComponent(name), {{ method: 'DELETE' }});
        const result = await res.json();
        if (result.ok) {{ closeModal(); location.reload(); }}
        else alert(result.error || 'Error deleting');
    }} catch(e) {{ alert('Error: ' + e.message); }}
}}

document.getElementById('editModal').addEventListener('click', function(e) {{
    if (e.target === this) closeModal();
}});

// Forecast charts (lazy init)
function initForecastCharts() {{
    window._forecastInit = true;
    const ctx = document.getElementById('monthlyChart');
    if (!ctx) return;
    new Chart(ctx, {{
        type: 'bar',
        data: {{
            labels: {all_months},
            datasets: [
                {{
                    label: 'Won Premium',
                    data: {monthly_rev_values},
                    backgroundColor: '#10b981',
                    borderRadius: 6,
                }},
                {{
                    label: 'Monthly Target',
                    data: {monthly_target_line},
                    type: 'line',
                    borderColor: '#ef4444',
                    borderDash: [5, 5],
                    borderWidth: 2,
                    pointRadius: 0,
                    fill: false,
                }}
            ]
        }},
        options: {{
            responsive: true,
            scales: {{
                y: {{
                    beginAtZero: true,
                    ticks: {{ callback: v => '$' + (v/1000).toFixed(0) + 'K' }}
                }}
            }},
            plugins: {{
                legend: {{ position: 'bottom', labels: {{ boxWidth: 12, padding: 8, font: {{ size: 11 }} }} }}
            }}
        }}
    }});
}}

// Velocity chart (lazy init)
function initFunnelCharts() {{
    window._funnelInit = true;
    const ctx = document.getElementById('velocityChart');
    if (!ctx) return;
    const velocityLabels = {list(avg_stage_days.keys())};
    const velocityData = {[round(v, 1) for v in avg_stage_days.values()]};
    new Chart(ctx, {{
        type: 'bar',
        data: {{
            labels: velocityLabels,
            datasets: [{{
                label: 'Avg Days in Stage',
                data: velocityData,
                backgroundColor: velocityData.map(d => d > 14 ? '#ef4444' : d > 7 ? '#f59e0b' : '#10b981'),
                borderRadius: 6,
            }}]
        }},
        options: {{
            indexAxis: 'y',
            responsive: true,
            scales: {{ x: {{ beginAtZero: true, title: {{ display: true, text: 'Days' }} }} }},
            plugins: {{ legend: {{ display: false }} }}
        }}
    }});
}}
</script>

</body>
</html>"""

    return Response(html, mimetype="text/html")


def run_dashboard():
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)


def start_dashboard_thread():
    t = threading.Thread(target=run_dashboard, daemon=True)
    t.start()
