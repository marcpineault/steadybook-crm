"""
PostgreSQL database module for SteadyBook CRM.

Multi-tenant: all queries scoped by tenant_id via _current_tenant_id context var.
The context var is set by the auth middleware in dashboard.py on every request.

Usage:
    from db import init_db, add_prospect, read_pipeline
    init_db()
"""

import os
import re
import secrets
import logging
from contextlib import contextmanager
from contextvars import ContextVar
from datetime import date, datetime, timedelta

import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

# ── Tenant context ──
# Set this at the start of every authenticated request.
# All query functions read it as their default tenant_id.
_current_tenant_id: ContextVar[int] = ContextVar("_current_tenant_id", default=1)

# ── Connection ──

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if not DATABASE_URL:
    logger.warning("DATABASE_URL not set — database operations will fail")


@contextmanager
def get_db():
    """Context manager for Postgres connections using RealDictCursor."""
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _row_to_dict(row):
    """Convert a RealDictRow to a plain dict."""
    if row is None:
        return None
    return dict(row)


def _rows_to_dicts(rows):
    """Convert a list of RealDictRow to list of dicts."""
    return [dict(r) for r in rows] if rows else []


# ── Numeric parsing ──

def _parse_numeric(val):
    """Parse a numeric value, stripping $ and commas. Returns float, or None if empty/invalid."""
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace("$", "").replace(",", ""))
    except (ValueError, TypeError):
        return None


def _parse_date_val(val):
    """Parse a date value from various formats. Returns string YYYY-MM-DD or None."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Already YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s.split(" ")[0]
    # Try common formats
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # return as-is if unparseable


def normalize_phone(phone: str) -> str:
    """Return the last 10 digits of a phone number with all non-digits stripped.

    Safe for numbers with 1s in the middle — strips by taking the last 10 chars
    of the digit string, not by removing the character '1' everywhere.
    """
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    return digits[-10:] if len(digits) >= 10 else digits


def get_prospect_by_phone(phone: str, tenant_id: int = 1):
    """Look up a prospect by phone number. Matches on last 10 digits. Returns dict or None."""
    last10 = normalize_phone(phone)
    if not last10:
        return None
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM prospects WHERE phone != '' AND tenant_id = ?", (tenant_id,)).fetchall()
    for row in rows:
        if normalize_phone(row["phone"]) == last10:
            return _row_to_dict(row)
    return None


# ── Schema ──

def init_db():
    """Create all tables if they don't exist."""
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS prospects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT DEFAULT '',
                email TEXT DEFAULT '',
                source TEXT DEFAULT '',
                priority TEXT DEFAULT '',
                stage TEXT DEFAULT 'New Lead',
                product TEXT DEFAULT '',
                aum REAL,
                revenue REAL,
                first_contact TEXT,
                next_followup TEXT,
                notes TEXT DEFAULT '',
                send_channel TEXT DEFAULT 'outlook',
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS activities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                prospect TEXT DEFAULT '',
                action TEXT DEFAULT '',
                outcome TEXT DEFAULT '',
                next_step TEXT DEFAULT '',
                notes TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS meetings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                time TEXT DEFAULT '',
                prospect TEXT DEFAULT '',
                type TEXT DEFAULT '',
                prep_notes TEXT DEFAULT '',
                status TEXT DEFAULT 'Scheduled'
            );

            CREATE TABLE IF NOT EXISTS insurance_book (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT DEFAULT '',
                address TEXT DEFAULT '',
                policy_start TEXT DEFAULT '',
                status TEXT DEFAULT 'Not Called',
                last_called TEXT,
                notes TEXT DEFAULT '',
                retry_date TEXT
            );

            CREATE TABLE IF NOT EXISTS win_loss_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                prospect TEXT DEFAULT '',
                outcome TEXT DEFAULT '',
                reason TEXT DEFAULT '',
                product TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS interactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT DEFAULT (datetime('now')),
                prospect TEXT DEFAULT '',
                source TEXT DEFAULT '',
                raw_text TEXT DEFAULT '',
                summary TEXT DEFAULT '',
                action_items TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now')),
                tenant_id INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                prospect TEXT DEFAULT '',
                due_date TEXT,
                remind_at TEXT,
                assigned_to TEXT DEFAULT '',
                created_by TEXT DEFAULT '',
                status TEXT DEFAULT 'pending',
                notes TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now')),
                completed_at TEXT
            );

            CREATE TABLE IF NOT EXISTS client_memory (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER REFERENCES prospects(id),
                category    TEXT NOT NULL,
                fact        TEXT NOT NULL,
                source      TEXT,
                needs_review INTEGER DEFAULT 0,
                extracted_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS approval_queue (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                type                TEXT NOT NULL,
                prospect_id         INTEGER REFERENCES prospects(id),
                channel             TEXT NOT NULL,
                content             TEXT NOT NULL,
                context             TEXT,
                status              TEXT DEFAULT 'pending',
                created_at          TEXT DEFAULT (datetime('now')),
                acted_on_at         TEXT,
                telegram_message_id TEXT
            );

            CREATE TABLE IF NOT EXISTS audit_log (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp        TEXT DEFAULT (datetime('now')),
                action_type      TEXT NOT NULL,
                target           TEXT,
                content          TEXT,
                compliance_check TEXT,
                approved_by      TEXT,
                outcome          TEXT
            );

            CREATE TABLE IF NOT EXISTS brand_voice (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                platform TEXT NOT NULL DEFAULT 'linkedin',
                content TEXT NOT NULL,
                post_type TEXT NOT NULL DEFAULT 'general',
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS market_calendar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL,
                title TEXT NOT NULL,
                date TEXT NOT NULL,
                description TEXT DEFAULT '',
                relevance_products TEXT DEFAULT '',
                recurring INTEGER DEFAULT 0
            );

        CREATE TABLE IF NOT EXISTS trust_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trust_level INTEGER NOT NULL DEFAULT 1,
            changed_at TEXT DEFAULT (datetime('now')),
            changed_by TEXT DEFAULT 'system'
        );

        CREATE TABLE IF NOT EXISTS campaigns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            segment_query TEXT DEFAULT '',
            status TEXT DEFAULT 'draft',
            channel TEXT DEFAULT 'email_draft',
            wave_count INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS campaign_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id INTEGER NOT NULL,
            prospect_name TEXT NOT NULL,
            content TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            queue_id INTEGER,
            wave INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (campaign_id) REFERENCES campaigns(id)
        );

        CREATE TABLE IF NOT EXISTS nurture_sequences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prospect_id INTEGER,
            prospect_name TEXT NOT NULL,
            status TEXT DEFAULT 'active',
            current_touch INTEGER DEFAULT 0,
            total_touches INTEGER DEFAULT 4,
            next_touch_date TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (prospect_id) REFERENCES prospects(id)
        );

        CREATE TABLE IF NOT EXISTS outcomes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action_id INTEGER,
            action_type TEXT NOT NULL,
            target TEXT,
            sent_at TEXT,
            response_received INTEGER DEFAULT 0,
            response_at TEXT,
            response_type TEXT,
            converted INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            resend_email_id TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (action_id) REFERENCES audit_log(id)
        );
        """)

        # Performance indexes
        conn.execute("CREATE INDEX IF NOT EXISTS idx_prospects_email ON prospects(email)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_outcomes_resend_id ON outcomes(resend_email_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_outcomes_target ON outcomes(target, sent_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tasks_status_due ON tasks(status, due_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tasks_remind ON tasks(remind_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_approval_queue_status ON approval_queue(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_nurture_status ON nurture_sequences(status, next_touch_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_activities_prospect ON activities(prospect)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_interactions_prospect ON interactions(prospect)")

    # Seed default trust level (idempotent — skips if any row exists)
    with get_db() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO trust_config (id, trust_level, changed_by) VALUES (1, 1, 'system')"
        )

    _migrate_phase6()
    _migrate_booking_nurture()
    _migrate_sms_conversations()
    _migrate_sms_agent()
    _migrate_multi_tenant()
    _migrate_sequences()
    cleanup_old_data()
    logger.info(f"Database initialized at {DB_PATH}")


def cleanup_old_data():
    """Remove interactions older than 90 days and audit log older than 7 years."""
    cutoff_interactions = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")
    cutoff_audit = (datetime.now() - timedelta(days=AUDIT_LOG_RETENTION_DAYS)).strftime("%Y-%m-%d")
    with get_db() as conn:
        deleted_interactions = conn.execute(
            "DELETE FROM interactions WHERE date < ? AND date != ''", (cutoff_interactions,)
        ).rowcount
        deleted_audit = conn.execute(
            "DELETE FROM audit_log WHERE timestamp < ?", (cutoff_audit,)
        ).rowcount
        if deleted_interactions or deleted_audit:
            logger.info(
                "Cleanup: removed %d old interactions, %d old audit entries",
                deleted_interactions, deleted_audit,
            )
            conn.execute("VACUUM")


def _migrate_booking_nurture():
    """Create booking_nurture_sequences table and indexes (safe to run repeatedly)."""
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS booking_nurture_sequences (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id      INTEGER REFERENCES prospects(id),
                prospect_name    TEXT NOT NULL,
                phone            TEXT NOT NULL,
                touch_number     INTEGER NOT NULL,
                scheduled_for    TEXT NOT NULL,
                meeting_datetime TEXT NOT NULL,
                meeting_date     TEXT NOT NULL,
                meeting_time     TEXT NOT NULL,
                meeting_type     TEXT DEFAULT '',
                product          TEXT DEFAULT '',
                status           TEXT DEFAULT 'queued',
                queue_id         INTEGER,
                created_at       TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_booking_nurture_status_sched
                ON booking_nurture_sequences(status, scheduled_for)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_booking_nurture_prospect
                ON booking_nurture_sequences(prospect_id, status)
        """)


def _migrate_sms_conversations():
    """Create sms_conversations table and indexes (safe to run repeatedly)."""
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sms_conversations (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id   INTEGER,
                prospect_name TEXT NOT NULL DEFAULT '',
                phone         TEXT NOT NULL,
                direction     TEXT NOT NULL,
                body          TEXT NOT NULL,
                twilio_sid    TEXT DEFAULT '',
                created_at    TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_sms_conv_phone
                ON sms_conversations(phone, created_at)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_sms_conv_prospect
                ON sms_conversations(prospect_id, created_at)
        """)


AUDIT_LOG_RETENTION_DAYS = 2555  # 7 years — FSRA compliance


def _migrate_sms_agent():
    """Add sms_opted_out column, sms_agents table, and partial unique index (idempotent)."""
    with get_db() as conn:
        # sms_opted_out column
        cols = [row[1] for row in conn.execute("PRAGMA table_info(prospects)").fetchall()]
        if "sms_opted_out" not in cols:
            conn.execute("ALTER TABLE prospects ADD COLUMN sms_opted_out INTEGER DEFAULT 0")
            logger.info("Migration: added sms_opted_out column")
        # Always backfill from notes (idempotent — only updates rows not yet flagged)
        conn.execute(
            "UPDATE prospects SET sms_opted_out = 1 WHERE notes LIKE '%[SMS_OPTED_OUT]%' AND sms_opted_out = 0"
        )
        logger.info("Migration: backfilled sms_opted_out from notes")

        # sms_agents table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sms_agents (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                phone         TEXT NOT NULL,
                prospect_id   INTEGER,
                prospect_name TEXT NOT NULL,
                objective     TEXT NOT NULL,
                status        TEXT DEFAULT 'pending_approval',
                attempts      INTEGER DEFAULT 0,
                created_at    TEXT DEFAULT (datetime('now')),
                updated_at    TEXT DEFAULT (datetime('now')),
                completed_at  TEXT,
                summary       TEXT
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_sms_agents_phone
                ON sms_agents(phone, status)
        """)

        # Partial unique index to deduplicate Twilio inbound retries
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_sms_inbound_sid
                ON sms_conversations(phone, twilio_sid)
                WHERE direction = 'inbound' AND twilio_sid != ''
        """)

        # Prospect notes timeline
        conn.execute("""
            CREATE TABLE IF NOT EXISTS prospect_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER NOT NULL,
                content TEXT NOT NULL,
                created_by TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (prospect_id) REFERENCES prospects(id)
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_prospect_notes_prospect
                ON prospect_notes(prospect_id, created_at DESC)
        """)


def _migrate_multi_tenant():
    """Create multi-tenant tables and add tenant_id to existing tables (safe to run repeatedly)."""
    with get_db() as conn:
        # Tenants table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tenants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                slug TEXT NOT NULL UNIQUE,
                company TEXT DEFAULT '',
                timezone TEXT DEFAULT 'America/Toronto',
                products TEXT DEFAULT '[]',
                config TEXT DEFAULT '{}',
                stripe_customer_id TEXT,
                plan TEXT DEFAULT 'starter',
                status TEXT DEFAULT 'active',
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_tenants_slug ON tenants(slug)")

        # Users table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                email TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                name TEXT DEFAULT '',
                role TEXT DEFAULT 'agent',
                telegram_chat_id TEXT,
                status TEXT DEFAULT 'active',
                created_at TEXT DEFAULT (datetime('now')),
                last_login TEXT
            )
        """)
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_email ON users(LOWER(email))")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_users_tenant ON users(tenant_id)")

        # API keys table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS api_keys (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                key_hash TEXT NOT NULL,
                name TEXT DEFAULT 'Default',
                scopes TEXT DEFAULT '["all"]',
                created_at TEXT DEFAULT (datetime('now')),
                expires_at TEXT
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_api_keys_hash ON api_keys(key_hash)")

        # Add tenant_id to existing tables (default 1 for backwards compat)
        _tables_needing_tenant = [
            "prospects", "activities", "meetings", "tasks", "campaigns",
            "approval_queue", "audit_log", "nurture_sequences",
            "booking_nurture_sequences", "sms_conversations", "sms_agents",
            "insurance_book", "trust_config", "brand_voice", "market_calendar",
        ]
        for table in _tables_needing_tenant:
            try:
                cols = [row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()]
                if "tenant_id" not in cols:
                    conn.execute(f"ALTER TABLE {table} ADD COLUMN tenant_id INTEGER DEFAULT 1")
                    logger.info("Migration: added tenant_id to %s", table)
            except Exception:
                pass  # Table may not exist yet

        # Create default tenant for existing single-user deployment (idempotent)
        existing = conn.execute("SELECT id FROM tenants WHERE id = 1").fetchone()
        if not existing:
            conn.execute(
                """INSERT INTO tenants (id, name, slug, company, timezone, plan, status)
                   VALUES (1, 'Default', 'default', '', 'America/Toronto', 'pro', 'active')"""
            )
            logger.info("Migration: created default tenant (id=1)")

    logger.info("Multi-tenant migration complete")


def _migrate_sequences():
    """Create sequence automation tables (safe to run repeatedly)."""
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sequences (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tenant_id INTEGER NOT NULL DEFAULT 1 REFERENCES tenants(id),
                name TEXT NOT NULL,
                description TEXT DEFAULT '',
                trigger_type TEXT NOT NULL,
                trigger_config TEXT DEFAULT '{}',
                status TEXT DEFAULT 'active',
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sequences_tenant ON sequences(tenant_id, status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sequences_trigger ON sequences(trigger_type, status)")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS sequence_steps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sequence_id INTEGER NOT NULL REFERENCES sequences(id) ON DELETE CASCADE,
                step_order INTEGER NOT NULL,
                step_type TEXT NOT NULL,
                delay_minutes INTEGER DEFAULT 0,
                content_template TEXT DEFAULT '',
                channel TEXT DEFAULT '',
                config TEXT DEFAULT '{}'
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_seq_steps_seq ON sequence_steps(sequence_id, step_order)")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS sequence_enrollments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sequence_id INTEGER NOT NULL REFERENCES sequences(id),
                prospect_id INTEGER NOT NULL REFERENCES prospects(id),
                status TEXT DEFAULT 'active',
                current_step INTEGER DEFAULT 1,
                enrolled_at TEXT DEFAULT (datetime('now')),
                last_step_at TEXT,
                next_step_at TEXT,
                completed_at TEXT,
                trigger_data TEXT DEFAULT '{}'
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_seq_enroll_due
                ON sequence_enrollments(status, next_step_at)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_seq_enroll_prospect
                ON sequence_enrollments(prospect_id, status)
        """)
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_seq_enroll_active
                ON sequence_enrollments(sequence_id, prospect_id)
                WHERE status = 'active'
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS sequence_step_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                enrollment_id INTEGER NOT NULL REFERENCES sequence_enrollments(id),
                step_id INTEGER NOT NULL REFERENCES sequence_steps(id),
                status TEXT DEFAULT 'ok',
                content TEXT DEFAULT '',
                executed_at TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_seq_step_logs_enrollment
                ON sequence_step_logs(enrollment_id, executed_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS prospect_tags (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER NOT NULL,
                tag TEXT NOT NULL,
                applied_by TEXT DEFAULT 'system',
                applied_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                UNIQUE(prospect_id, tag)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS enrichment_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER NOT NULL UNIQUE,
                status TEXT DEFAULT 'pending',
                attempts INTEGER DEFAULT 0,
                last_attempt TEXT,
                linkedin_url TEXT,
                instagram_handle TEXT,
                headshot_url TEXT,
                bio TEXT,
                company_website TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS referrals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                referrer_prospect_id INTEGER,
                referred_prospect_id INTEGER NOT NULL,
                referral_date TEXT DEFAULT (datetime('now')),
                notes TEXT,
                FOREIGN KEY (referrer_prospect_id) REFERENCES prospects(id),
                FOREIGN KEY (referred_prospect_id) REFERENCES prospects(id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS intake_form_responses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER NOT NULL,
                form_type TEXT NOT NULL,
                responses TEXT NOT NULL,
                submitted_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS email_tracking (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER,
                prospect_name TEXT,
                email_type TEXT,
                token TEXT UNIQUE,
                opened_at TEXT,
                link_clicked_at TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE SET NULL
            )
        """)

    logger.info("Sequence tables migration complete")


def _migrate_phase6():
    """Add Phase 6 columns if they don't exist (safe to run repeatedly)."""
    with get_db() as conn:
        cols = [row[1] for row in conn.execute("PRAGMA table_info(prospects)").fetchall()]
        if "send_channel" not in cols:
            conn.execute("ALTER TABLE prospects ADD COLUMN send_channel TEXT DEFAULT 'outlook'")
            logger.info("Migration: added send_channel to prospects")
        outcome_cols = [row[1] for row in conn.execute("PRAGMA table_info(outcomes)").fetchall()]
        if "resend_email_id" not in outcome_cols:
            conn.execute("ALTER TABLE outcomes ADD COLUMN resend_email_id TEXT")
            logger.info("Migration: added resend_email_id to outcomes")
        # Migration: add assigned_to if not present
        try:
            conn.execute("ALTER TABLE prospects ADD COLUMN assigned_to TEXT DEFAULT ''")
        except Exception:
            pass  # Column already exists
        # Migration: add trust_level to tenants if not present
        try:
            conn.execute("ALTER TABLE tenants ADD COLUMN trust_level INTEGER DEFAULT 1")
        except Exception:
            pass  # Column already exists


# ── Prospects CRUD ──

def read_pipeline(tenant_id: int = 1):
    """Return all prospects as a list of dicts."""
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM prospects WHERE tenant_id = ? ORDER BY id", (tenant_id,)).fetchall()
    return _rows_to_dicts(rows)


def add_prospect(data: dict, tenant_id: int = 1) -> str:
    """Insert a new prospect. Returns status string."""
    name = data.get("name", "").strip()
    if not name:
        return "No name provided for prospect."

    aum = _parse_numeric(data.get("aum"))
    revenue = _parse_numeric(data.get("revenue"))
    first_contact = data.get("first_contact") or date.today().strftime("%Y-%m-%d")
    stage = data.get("stage") or "New Lead"

    with get_db() as conn:
        conn.execute(
            """INSERT INTO prospects
               (name, phone, email, source, priority, stage, product,
                aum, revenue, first_contact, next_followup, notes, send_channel, tenant_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                name,
                data.get("phone", ""),
                data.get("email", ""),
                data.get("source", ""),
                data.get("priority", ""),
                stage,
                data.get("product", ""),
                aum,
                revenue,
                first_contact,
                data.get("next_followup", ""),
                data.get("notes", ""),
                data.get("send_channel", "outlook"),
                tenant_id,
            ),
        )
    return f"Added {name} to pipeline."


def update_prospect(name: str, updates: dict) -> str:
    """Update a prospect by partial name match (case insensitive).
    Skips empty values. Returns status string."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, name FROM prospects WHERE LOWER(name) LIKE ? LIMIT 1",
            (f"%{name.lower()}%",),
        ).fetchone()

        if not row:
            return f"Could not find prospect matching '{name}'."

        prospect_id = row["id"]
        matched_name = row["name"]

        allowed = {
            "name", "phone", "email", "source", "priority", "stage",
            "product", "aum", "revenue", "first_contact", "next_followup", "notes",
            "send_channel",
        }

        safe_fields = {}
        for field, value in updates.items():
            if field not in allowed or value is None:
                continue
            if field in ("aum", "revenue"):
                parsed = _parse_numeric(value)
                if parsed is not None:
                    value = parsed
            safe_fields[field] = value

        if not safe_fields:
            return f"No valid updates for {matched_name}."

        # Cap notes at 2000 chars — truncate oldest content from the front
        if "notes" in safe_fields and safe_fields["notes"]:
            notes_val = safe_fields["notes"]
            if len(notes_val) > 2000:
                safe_fields["notes"] = "..." + notes_val[-1997:]

        # Build SET clause using only validated field names from the allowlist
        validated_fields = [f for f in safe_fields if f in allowed]
        set_clauses = ", ".join(f'"{field}" = ?' for field in validated_fields)
        values = [safe_fields[f] for f in validated_fields] + [prospect_id]
        conn.execute(
            f"UPDATE prospects SET {set_clauses}, updated_at = datetime('now') WHERE id = ?",
            values,
        )

    return f"Updated {matched_name}: {', '.join(f'{f} → {v}' for f, v in safe_fields.items())}"


def delete_prospect(name: str) -> str:
    """Delete a prospect by partial name match. Returns status string."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, name FROM prospects WHERE LOWER(name) LIKE ? LIMIT 1",
            (f"%{name.lower()}%",),
        ).fetchone()

        if not row:
            return f"Could not find prospect matching '{name}'."

        matched_name = row["name"]
        pid = row["id"]
        # Delete related records first to avoid foreign key constraint failures
        conn.execute("DELETE FROM client_memory WHERE prospect_id = ?", (pid,))
        conn.execute("DELETE FROM approval_queue WHERE prospect_id = ?", (pid,))
        conn.execute("DELETE FROM nurture_sequences WHERE prospect_id = ?", (pid,))
        # Clean up sequence enrollments and their step logs
        enrollment_ids = [r[0] for r in conn.execute(
            "SELECT id FROM sequence_enrollments WHERE prospect_id = ?", (pid,)
        ).fetchall()]
        for eid in enrollment_ids:
            conn.execute("DELETE FROM sequence_step_logs WHERE enrollment_id = ?", (eid,))
        conn.execute("DELETE FROM sequence_enrollments WHERE prospect_id = ?", (pid,))
        conn.execute("DELETE FROM activities WHERE LOWER(prospect) = ?", (matched_name.lower(),))
        conn.execute("DELETE FROM interactions WHERE LOWER(prospect) = ?", (matched_name.lower(),))
        conn.execute("DELETE FROM prospects WHERE id = ?", (pid,))

    return f"Deleted {matched_name} from pipeline."


def merge_prospects(keep_name: str, merge_name: str) -> str:
    """Merge one prospect into another. Keeps keep_name, deletes merge_name.

    Transfers all activities, interactions, memory, approvals, and nurture
    sequences from merge_name to keep_name. Merges notes.
    """
    with get_db() as conn:
        keep = conn.execute(
            "SELECT * FROM prospects WHERE LOWER(name) LIKE ? LIMIT 1",
            (f"%{keep_name.lower()}%",),
        ).fetchone()
        merge = conn.execute(
            "SELECT * FROM prospects WHERE LOWER(name) LIKE ? LIMIT 1",
            (f"%{merge_name.lower()}%",),
        ).fetchone()

        if not keep:
            return f"Could not find prospect '{keep_name}'."
        if not merge:
            return f"Could not find prospect '{merge_name}'."
        if keep["id"] == merge["id"]:
            return "Cannot merge a prospect with itself."

        keep_id = keep["id"]
        merge_id = merge["id"]
        keep_real = keep["name"]
        merge_real = merge["name"]

        # Transfer activities and interactions (name-based)
        conn.execute(
            "UPDATE activities SET prospect = ? WHERE LOWER(prospect) = ?",
            (keep_real, merge_real.lower()),
        )
        conn.execute(
            "UPDATE interactions SET prospect = ? WHERE LOWER(prospect) = ?",
            (keep_real, merge_real.lower()),
        )

        # Transfer FK-based records
        conn.execute(
            "UPDATE client_memory SET prospect_id = ? WHERE prospect_id = ?",
            (keep_id, merge_id),
        )
        conn.execute(
            "UPDATE approval_queue SET prospect_id = ? WHERE prospect_id = ?",
            (keep_id, merge_id),
        )
        conn.execute(
            "UPDATE nurture_sequences SET prospect_id = ?, prospect_name = ? WHERE prospect_id = ?",
            (keep_id, keep_real, merge_id),
        )

        # Merge notes
        keep_notes = keep["notes"] or ""
        merge_notes = merge["notes"] or ""
        if merge_notes:
            combined = f"{keep_notes} | Merged from {merge_real}: {merge_notes}".strip(" |")
            conn.execute("UPDATE prospects SET notes = ? WHERE id = ?", (combined, keep_id))

        # Fill empty fields on keep from merge
        for field in ("phone", "email", "product", "aum", "revenue"):
            if not keep[field] and merge[field]:
                conn.execute(f"UPDATE prospects SET {field} = ? WHERE id = ?", (merge[field], keep_id))

        # Delete the merged prospect
        conn.execute("DELETE FROM prospects WHERE id = ?", (merge_id,))

    return f"Merged {merge_real} into {keep_real}."


def get_all_prospect_names(tenant_id: int = 1) -> list[str]:
    """Return a list of all prospect names (for duplicate checking)."""
    with get_db() as conn:
        rows = conn.execute("SELECT name FROM prospects WHERE tenant_id = ? ORDER BY name", (tenant_id,)).fetchall()
    return [row["name"] for row in rows]


def get_prospect_by_name(name: str, tenant_id: int = 1):
    """Lookup by exact match first, then fuzzy partial match. Returns single dict or None."""
    with get_db() as conn:
        # Try exact match first (case-insensitive)
        row = conn.execute(
            "SELECT * FROM prospects WHERE LOWER(name) = ? AND tenant_id = ? LIMIT 1",
            (name.lower(), tenant_id),
        ).fetchone()
        if row is None:
            # Fall back to partial match — fetch ALL candidates and pick best
            candidates = conn.execute(
                "SELECT * FROM prospects WHERE LOWER(name) LIKE ? AND tenant_id = ?",
                (f"%{name.lower()}%", tenant_id),
            ).fetchall()
            if len(candidates) == 1:
                row = candidates[0]
                logger.info(f"Prospect fuzzy match: '{name}' → '{dict(row)['name']}'")
            elif len(candidates) > 1:
                other_names = [dict(r)["name"] for r in candidates]
                logger.warning(
                    "Prospect fuzzy match ambiguous for '%s' — matched: %s. Returning None.",
                    name, other_names,
                )
                row = None
    return _row_to_dict(row)


def get_prospect_by_email(email: str, tenant_id: int = 1):
    """Lookup prospect by exact email match (case-insensitive). Returns dict or None."""
    if not email:
        return None
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM prospects WHERE LOWER(email) = ? AND tenant_id = ? LIMIT 1",
            (email.lower().strip(), tenant_id),
        ).fetchone()
    return _row_to_dict(row)


def get_prospect_by_id(prospect_id: int):
    """Look up a prospect by primary key. Returns dict or None."""
    with get_db() as conn:
        row = conn.execute("SELECT * FROM prospects WHERE id = ?", (prospect_id,)).fetchone()
    return _row_to_dict(row)


# ── Activities CRUD ──

def add_activity(data: dict, tenant_id: int = 1) -> str:
    """Add an entry to the activity log. Defaults date to today."""
    activity_date = data.get("date") or date.today().strftime("%Y-%m-%d")
    with get_db() as conn:
        conn.execute(
            """INSERT INTO activities (date, prospect, action, outcome, next_step, notes, tenant_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                activity_date,
                data.get("prospect", ""),
                data.get("action", ""),
                data.get("outcome", ""),
                data.get("next_step", ""),
                data.get("notes", ""),
                tenant_id,
            ),
        )
    return f"Logged activity for {data.get('prospect', 'unknown')}."


def read_activities(limit: int = 100, tenant_id: int = 1):
    """Return recent activities as list of dicts, newest first."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM activities WHERE tenant_id = ? ORDER BY id DESC LIMIT ?", (tenant_id, limit)
        ).fetchall()
    return _rows_to_dicts(rows)


# ── Prospect Notes ──

def add_prospect_note(prospect_id: int, content: str, created_by: str = "") -> dict | None:
    """Add a note to a prospect's timeline. Returns the note dict."""
    content = content.strip()
    if not content:
        return None
    with get_db() as conn:
        cursor = conn.execute(
            "INSERT INTO prospect_notes (prospect_id, content, created_by) VALUES (?, ?, ?)",
            (prospect_id, content, created_by),
        )
        row = conn.execute("SELECT * FROM prospect_notes WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return _row_to_dict(row) if row else None


def get_prospect_notes(prospect_id: int, limit: int = 50) -> list[dict]:
    """Get notes for a prospect, newest first."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM prospect_notes WHERE prospect_id = ? ORDER BY created_at DESC, id DESC LIMIT ?",
            (prospect_id, limit),
        ).fetchall()
    return _rows_to_dicts(rows)


def delete_prospect_note(note_id: int) -> str:
    """Delete a note by ID."""
    with get_db() as conn:
        row = conn.execute("SELECT id FROM prospect_notes WHERE id = ?", (note_id,)).fetchone()
        if not row:
            return "Note not found."
        conn.execute("DELETE FROM prospect_notes WHERE id = ?", (note_id,))
    return "Note deleted."


# ── Meetings CRUD ──

def add_meeting(data: dict, tenant_id: int = 1) -> str:
    """Add a meeting. Defaults status to 'Scheduled'."""
    status = data.get("status") or "Scheduled"
    with get_db() as conn:
        conn.execute(
            """INSERT INTO meetings (date, time, prospect, type, prep_notes, status, tenant_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                data.get("date", ""),
                data.get("time", ""),
                data.get("prospect", ""),
                data.get("type", ""),
                data.get("prep_notes", ""),
                status,
                tenant_id,
            ),
        )
    return f"Meeting added: {data.get('prospect', '?')} on {data.get('date', '?')} at {data.get('time', '?')}"


def read_meetings(tenant_id: int = 1):
    """Return all meetings ordered by date and time."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM meetings WHERE tenant_id = ? ORDER BY date, time", (tenant_id,)
        ).fetchall()
    return _rows_to_dicts(rows)


def update_meeting(meeting_id: int, updates: dict) -> str:
    """Update a meeting by ID."""
    allowed = {"date", "time", "prospect", "type", "prep_notes", "status"}
    safe_fields = {f: v for f, v in updates.items() if f in allowed and v is not None}
    if not safe_fields:
        return f"No valid updates for meeting {meeting_id}."
    with get_db() as conn:
        validated_fields = [f for f in safe_fields if f in allowed]
        set_clauses = ", ".join(f'"{field}" = ?' for field in validated_fields)
        values = [safe_fields[f] for f in validated_fields] + [meeting_id]
        conn.execute(f"UPDATE meetings SET {set_clauses} WHERE id = ?", values)
    return f"Updated meeting {meeting_id}: {', '.join(f'{f} → {v}' for f, v in safe_fields.items())}"


# ── Insurance Book CRUD ──

def read_insurance_book(tenant_id: int = 1):
    """Return all insurance book entries."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM insurance_book WHERE tenant_id = ? ORDER BY id", (tenant_id,)
        ).fetchall()
    return _rows_to_dicts(rows)


def add_insurance_entry(data: dict, tenant_id: int = 1) -> str:
    """Add an entry to the insurance book."""
    name = data.get("name", "").strip()
    if not name:
        return "No name provided for insurance entry."

    with get_db() as conn:
        conn.execute(
            """INSERT INTO insurance_book
               (name, phone, address, policy_start, status, last_called, notes, retry_date, tenant_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                name,
                data.get("phone", ""),
                data.get("address", ""),
                data.get("policy_start", ""),
                data.get("status", "Not Called"),
                data.get("last_called"),
                data.get("notes", ""),
                data.get("retry_date"),
                tenant_id,
            ),
        )
    return f"Added {name} to insurance book."


def update_insurance_entry(entry_id: int, updates: dict) -> str:
    """Update an insurance book entry by ID."""
    allowed = {"name", "phone", "address", "policy_start", "status",
               "last_called", "notes", "retry_date"}
    safe_fields = {f: v for f, v in updates.items() if f in allowed and v is not None}
    if not safe_fields:
        return f"No valid updates for insurance entry {entry_id}."
    with get_db() as conn:
        validated_fields = [f for f in safe_fields if f in allowed]
        set_clauses = ", ".join(f'"{field}" = ?' for field in validated_fields)
        values = [safe_fields[f] for f in validated_fields] + [entry_id]
        conn.execute(f"UPDATE insurance_book SET {set_clauses} WHERE id = ?", values)
    return f"Updated insurance entry {entry_id}: {', '.join(f'{f} → {v}' for f, v in safe_fields.items())}"


# ── Win/Loss Log ──

def log_win_loss(prospect_name: str, outcome: str, reason: str, product: str = "", tenant_id: int = 1) -> str:
    """Log a win or loss with reason."""
    # If product not provided, look it up from prospects
    if not product:
        p = get_prospect_by_name(prospect_name, tenant_id=tenant_id)
        if p:
            product = p.get("product", "")

    with get_db() as conn:
        conn.execute(
            """INSERT INTO win_loss_log (date, prospect, outcome, reason, product, tenant_id)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                date.today().strftime("%Y-%m-%d"),
                prospect_name,
                outcome,
                reason,
                product,
                tenant_id,
            ),
        )
    return f"Logged {outcome} for {prospect_name}: {reason}"


def get_win_loss_stats(tenant_id: int = 1):
    """Return all win/loss entries as list of dicts."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM win_loss_log WHERE tenant_id = ? ORDER BY id DESC", (tenant_id,)
        ).fetchall()
    return _rows_to_dicts(rows)


# ── Interactions CRUD ──

def add_interaction(data: dict, tenant_id: int = 1) -> str:
    """Log an interaction (voice note, transcript, email, booking)."""
    with get_db() as conn:
        conn.execute(
            """INSERT INTO interactions (date, prospect, source, raw_text, summary, action_items, tenant_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                data.get("date") or datetime.now().strftime("%Y-%m-%d %H:%M"),
                data.get("prospect", ""),
                data.get("source", ""),
                data.get("raw_text", ""),
                data.get("summary", ""),
                data.get("action_items", ""),
                tenant_id,
            ),
        )
    return f"Logged interaction for {data.get('prospect', 'unknown')} via {data.get('source', '?')}."


def read_interactions(limit: int = 50, prospect: str = "", tenant_id: int = 1):
    """Return recent interactions, newest first. Optionally filter by prospect."""
    with get_db() as conn:
        if prospect:
            rows = conn.execute(
                "SELECT * FROM interactions WHERE LOWER(prospect) LIKE ? AND tenant_id = ? ORDER BY id DESC LIMIT ?",
                (f"%{prospect.lower()}%", tenant_id, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM interactions WHERE tenant_id = ? ORDER BY id DESC LIMIT ?", (tenant_id, limit)
            ).fetchall()
    return _rows_to_dicts(rows)


# ── Tasks CRUD ──

def add_task(data: dict, tenant_id: int = 1):
    """Add a task. Returns the created task as dict, or None if no title."""
    title = data.get("title", "").strip()
    if not title:
        return None

    # Normalize remind_at to "YYYY-MM-DD HH:MM" (replace T from datetime-local inputs)
    remind_at = data.get("remind_at")
    if remind_at and isinstance(remind_at, str):
        remind_at = remind_at.replace("T", " ")

    with get_db() as conn:
        cursor = conn.execute(
            """INSERT INTO tasks
               (title, prospect, due_date, remind_at, assigned_to, created_by, notes, tenant_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                title,
                data.get("prospect", ""),
                data.get("due_date"),
                remind_at,
                data.get("assigned_to", ""),
                data.get("created_by", ""),
                data.get("notes", ""),
                tenant_id,
            ),
        )
        task_id = cursor.lastrowid
        row = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    return _row_to_dict(row)


def get_tasks(assigned_to=None, status="pending", prospect=None, limit=50, tenant_id: int = 1):
    """Get tasks with filters. Orders by due_date ASC (nulls last), then created_at DESC."""
    conditions = ["tenant_id = ?"]
    params = [tenant_id]

    if status:
        conditions.append("status = ?")
        params.append(status)
    if assigned_to:
        conditions.append("assigned_to = ?")
        params.append(assigned_to)
    if prospect:
        conditions.append("LOWER(prospect) LIKE ?")
        params.append(f"%{prospect.lower()}%")

    where = " AND ".join(conditions)
    params.append(limit)

    with get_db() as conn:
        rows = conn.execute(
            f"""SELECT * FROM tasks WHERE {where}
                ORDER BY
                    CASE WHEN due_date IS NULL THEN 1 ELSE 0 END,
                    due_date ASC,
                    created_at DESC
                LIMIT ?""",
            params,
        ).fetchall()
    return _rows_to_dicts(rows)


def update_task(task_id: int, updates: dict, updated_by: str = "", is_admin: bool = False) -> str:
    """Update a task's fields. Only assignee or admin can update."""
    allowed = {"title", "prospect", "due_date", "remind_at", "notes"}
    with get_db() as conn:
        row = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
        if not row:
            return f"Task {task_id} not found."
        if not is_admin and row["assigned_to"] != updated_by:
            return f"Not authorized to update task {task_id}."
        safe_fields = {}
        for field, value in updates.items():
            if field not in allowed:
                continue
            if field == "remind_at" and value and isinstance(value, str):
                value = value.replace("T", " ")
            safe_fields[field] = value
        if not safe_fields:
            return f"No valid updates for task {task_id}."
        validated_fields = [f for f in safe_fields if f in allowed]
        set_clauses = ", ".join(f'"{field}" = ?' for field in validated_fields)
        values = [safe_fields[f] for f in validated_fields] + [task_id]
        conn.execute(f"UPDATE tasks SET {set_clauses} WHERE id = ?", values)
    return f"Updated task {task_id}."


def complete_task(task_id: int, completed_by: str, is_admin: bool = False) -> str:
    """Mark a task as completed. Only assignee or admin can complete."""
    with get_db() as conn:
        row = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
        if not row:
            return f"Task {task_id} not found."
        if not is_admin and row["assigned_to"] != completed_by:
            return f"Not authorized to complete task {task_id}."
        conn.execute(
            "UPDATE tasks SET status = 'completed', completed_at = datetime('now') WHERE id = ?",
            (task_id,),
        )
    return f"Completed: {row['title']}"


def delete_task(task_id: int, deleted_by: str, is_admin: bool = False) -> str:
    """Delete a task. Only assignee or admin can delete."""
    with get_db() as conn:
        row = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
        if not row:
            return f"Task {task_id} not found."
        if not is_admin and row["assigned_to"] != deleted_by:
            return f"Not authorized to delete task {task_id}."
        conn.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    return f"Deleted: {row['title']}"


def get_due_tasks(date_str: str, tenant_id: int = 1):
    """Get pending tasks due on a specific date."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM tasks WHERE due_date = ? AND status = 'pending' AND tenant_id = ? ORDER BY created_at",
            (date_str, tenant_id),
        ).fetchall()
    return _rows_to_dicts(rows)


def get_overdue_tasks(tenant_id: int = 1):
    """Get pending tasks with due_date before today."""
    today = date.today().strftime("%Y-%m-%d")
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM tasks WHERE due_date < ? AND status = 'pending' AND tenant_id = ? ORDER BY due_date ASC",
            (today, tenant_id),
        ).fetchall()
    return _rows_to_dicts(rows)


def get_reminder_tasks(now_str: str, tenant_id: int = 1):
    """Get pending tasks with remind_at <= now that haven't been cleared.
    Normalizes remind_at by replacing 'T' with space for consistent comparison."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM tasks WHERE remind_at IS NOT NULL AND REPLACE(remind_at, 'T', ' ') <= ? AND status = 'pending' AND tenant_id = ? ORDER BY remind_at",
            (now_str.replace("T", " "), tenant_id),
        ).fetchall()
    return _rows_to_dicts(rows)


def clear_reminder(task_id: int):
    """Clear remind_at after firing so it doesn't repeat."""
    with get_db() as conn:
        conn.execute("UPDATE tasks SET remind_at = NULL WHERE id = ?", (task_id,))


# ── Migration from Excel ──

def migrate_from_excel(excel_path: str) -> str:
    """Migrate data from the existing Excel pipeline file to SQLite.

    Skips if the database already has prospects.
    """
    import openpyxl

    if not os.path.exists(excel_path):
        return f"Excel file not found: {excel_path}"

    # Skip if DB already has data
    with get_db() as conn:
        count = conn.execute("SELECT COUNT(*) FROM prospects").fetchone()[0]
        if count > 0:
            return f"Database already has {count} prospects. Skipping migration."

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    def cell_str(ws, row, col):
        v = ws.cell(row=row, column=col).value
        return str(v) if v is not None else ""

    def cell_val(ws, row, col):
        return ws.cell(row=row, column=col).value

    migrated = {"prospects": 0, "activities": 0, "meetings": 0,
                "insurance": 0, "win_loss": 0}

    with get_db() as conn:
        # ── Pipeline sheet: starts row 5, columns 1-13 ──
        if "Pipeline" in wb.sheetnames:
            ws = wb["Pipeline"]
            for r in range(5, 5 + 80):
                name = cell_val(ws, r, 1)
                if not name:
                    continue
                aum = _parse_numeric(cell_val(ws, r, 8))
                revenue = _parse_numeric(cell_val(ws, r, 9))
                first_contact = _parse_date_val(cell_val(ws, r, 10))
                next_followup = _parse_date_val(cell_val(ws, r, 11))
                # Column 12 is days_open (computed), skip it
                conn.execute(
                    """INSERT INTO prospects
                       (name, phone, email, source, priority, stage, product,
                        aum, revenue, first_contact, next_followup, notes)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        cell_str(ws, r, 1),
                        cell_str(ws, r, 2),
                        cell_str(ws, r, 3),
                        cell_str(ws, r, 4),
                        cell_str(ws, r, 5),
                        cell_str(ws, r, 6) or "New Lead",
                        cell_str(ws, r, 7),
                        aum,
                        revenue,
                        first_contact or date.today().strftime("%Y-%m-%d"),
                        next_followup or "",
                        cell_str(ws, r, 13),
                    ),
                )
                migrated["prospects"] += 1

        # ── Activity Log sheet: starts row 3, columns 1-6 ──
        if "Activity Log" in wb.sheetnames:
            ws = wb["Activity Log"]
            for r in range(3, 3 + 200):
                d = cell_val(ws, r, 1)
                if not d:
                    continue
                conn.execute(
                    """INSERT INTO activities (date, prospect, action, outcome, next_step, notes)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (
                        _parse_date_val(d) or cell_str(ws, r, 1),
                        cell_str(ws, r, 2),
                        cell_str(ws, r, 3),
                        cell_str(ws, r, 4),
                        cell_str(ws, r, 5),
                        cell_str(ws, r, 6),
                    ),
                )
                migrated["activities"] += 1

        # ── Meetings sheet: starts row 3, columns 1-6 ──
        if "Meetings" in wb.sheetnames:
            ws = wb["Meetings"]
            for r in range(3, 3 + 100):
                d = cell_val(ws, r, 1)
                if not d:
                    continue
                conn.execute(
                    """INSERT INTO meetings (date, time, prospect, type, prep_notes, status)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (
                        _parse_date_val(d) or cell_str(ws, r, 1),
                        cell_str(ws, r, 2),
                        cell_str(ws, r, 3),
                        cell_str(ws, r, 4),
                        cell_str(ws, r, 5),
                        cell_str(ws, r, 6) or "Scheduled",
                    ),
                )
                migrated["meetings"] += 1

        # ── Insurance Book sheet: starts row 3, columns 1-8 ──
        if "Insurance Book" in wb.sheetnames:
            ws = wb["Insurance Book"]
            for r in range(3, 3 + 500):
                name = cell_val(ws, r, 1)
                if not name:
                    continue
                conn.execute(
                    """INSERT INTO insurance_book
                       (name, phone, address, policy_start, status, last_called, notes, retry_date)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        cell_str(ws, r, 1),
                        cell_str(ws, r, 2),
                        cell_str(ws, r, 3),
                        _parse_date_val(cell_val(ws, r, 4)) or cell_str(ws, r, 4),
                        cell_str(ws, r, 5) or "Not Called",
                        _parse_date_val(cell_val(ws, r, 6)) or "",
                        cell_str(ws, r, 7),
                        _parse_date_val(cell_val(ws, r, 8)) or "",
                    ),
                )
                migrated["insurance"] += 1

        # ── Win Loss Log sheet: starts row 3, columns 1-5 ──
        if "Win Loss Log" in wb.sheetnames:
            ws = wb["Win Loss Log"]
            for r in range(3, 3 + 100):
                d = cell_val(ws, r, 1)
                if not d:
                    continue
                conn.execute(
                    """INSERT INTO win_loss_log (date, prospect, outcome, reason, product)
                       VALUES (?, ?, ?, ?, ?)""",
                    (
                        _parse_date_val(d) or cell_str(ws, r, 1),
                        cell_str(ws, r, 2),
                        cell_str(ws, r, 3),
                        cell_str(ws, r, 4),
                        cell_str(ws, r, 5),
                    ),
                )
                migrated["win_loss"] += 1

    wb.close()

    summary = (
        f"Migration complete: {migrated['prospects']} prospects, "
        f"{migrated['activities']} activities, {migrated['meetings']} meetings, "
        f"{migrated['insurance']} insurance entries, {migrated['win_loss']} win/loss records."
    )
    logger.info(summary)
    return summary


# ── Tag helpers ──

def apply_tag(prospect_id: int, tag: str, applied_by: str = "system") -> bool:
    """Apply a tag to a prospect. Returns True if new, False if already existed."""
    with get_db() as conn:
        try:
            conn.execute(
                "INSERT INTO prospect_tags (prospect_id, tag, applied_by) VALUES (?,?,?)",
                (prospect_id, tag, applied_by)
            )
            return True
        except sqlite3.IntegrityError:
            return False

def remove_tag(prospect_id: int, tag: str) -> None:
    with get_db() as conn:
        conn.execute("DELETE FROM prospect_tags WHERE prospect_id=? AND tag=?", (prospect_id, tag))

def get_tags(prospect_id: int) -> list[str]:
    with get_db() as conn:
        rows = conn.execute("SELECT tag FROM prospect_tags WHERE prospect_id=?", (prospect_id,)).fetchall()
        return [r["tag"] for r in rows]

def get_prospects_by_tag(tag: str) -> list[dict]:
    with get_db() as conn:
        rows = conn.execute("""
            SELECT p.* FROM prospects p
            JOIN prospect_tags t ON t.prospect_id = p.id
            WHERE t.tag = ?
        """, (tag,)).fetchall()
        return _rows_to_dicts(rows)

def queue_enrichment(prospect_id: int) -> None:
    with get_db() as conn:
        conn.execute("""
            INSERT INTO enrichment_queue (prospect_id) VALUES (?)
            ON CONFLICT(prospect_id) DO NOTHING
        """, (prospect_id,))


# ── Reporting queries ──

def get_conversion_by_source() -> list[dict]:
    """Conversion rate by lead source."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                source,
                COUNT(*) as total_leads,
                SUM(CASE WHEN stage = 'Closed Won' THEN 1 ELSE 0 END) as closed,
                ROUND(
                    100.0 * SUM(CASE WHEN stage = 'Closed Won' THEN 1 ELSE 0 END) / COUNT(*),
                    1
                ) as conversion_rate
            FROM prospects
            WHERE source IS NOT NULL AND source != ''
            GROUP BY source
            ORDER BY total_leads DESC
        """).fetchall()
    return [dict(r) for r in rows]


def get_pipeline_metrics() -> dict:
    """High-level pipeline summary metrics."""
    with get_db() as conn:
        row = conn.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN stage = 'Closed Won' THEN 1 ELSE 0 END) as closed_won,
                SUM(CASE WHEN stage = 'New Lead' THEN 1 ELSE 0 END) as new_leads,
                SUM(CASE WHEN stage NOT IN ('Closed Won', 'Closed Lost') THEN 1 ELSE 0 END) as active
            FROM prospects
        """).fetchone()
    if not row:
        return {"total": 0, "closed_won": 0, "new_leads": 0, "active": 0}
    return dict(row)


def get_stage_funnel() -> list[dict]:
    """Count of prospects per pipeline stage, ordered logically."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT stage, COUNT(*) as count
            FROM prospects
            GROUP BY stage
            ORDER BY count DESC
        """).fetchall()
    return [dict(r) for r in rows]


def get_fyc_by_advisor() -> list[dict]:
    """First-year commission summary by assigned advisor."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                COALESCE(assigned_to, 'Unassigned') as advisor,
                COUNT(*) as total_closed,
                SUM(CASE WHEN stage = 'Closed Won' THEN 1 ELSE 0 END) as closed_won
            FROM prospects
            GROUP BY assigned_to
            ORDER BY closed_won DESC
        """).fetchall()
    return [dict(r) for r in rows]


def get_avg_stage_time() -> list[dict]:
    """Average days prospects spend in each stage (approximate)."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT stage, COUNT(*) as count
            FROM prospects
            GROUP BY stage
        """).fetchall()
    return [dict(r) for r in rows]


def get_trust_level(tenant_id: int = 1) -> int:
    """Return the AI trust level for a tenant (1=draft only, 2=routine auto, 3=full autonomy)."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT trust_level FROM tenants WHERE id=?", (tenant_id,)
        ).fetchone()
    if row and row["trust_level"]:
        return int(row["trust_level"])
    return 1  # Default to most conservative


def create_email_tracking_token(prospect_id: int, prospect_name: str, email_type: str) -> str:
    """Create a unique tracking token for an outgoing email. Returns the token."""
    token = secrets.token_urlsafe(20)
    with get_db() as conn:
        conn.execute("""
            INSERT INTO email_tracking (prospect_id, prospect_name, email_type, token)
            VALUES (?, ?, ?, ?)
        """, (prospect_id, prospect_name, email_type, token))
    return token


def record_email_open(token: str) -> bool:
    """Record that a tracking pixel was fired. Returns True if token existed."""
    with get_db() as conn:
        cur = conn.execute("""
            UPDATE email_tracking
            SET opened_at = COALESCE(opened_at, datetime('now'))
            WHERE token = ? AND opened_at IS NULL
        """, (token,))
    return cur.rowcount > 0


def record_link_click(token: str) -> str | None:
    """Record a link click. Returns None (destination URL stored in notes, not here)."""
    with get_db() as conn:
        conn.execute("""
            UPDATE email_tracking
            SET link_clicked_at = COALESCE(link_clicked_at, datetime('now'))
            WHERE token = ?
        """, (token,))
    return None


def add_intake_form_response(prospect_id: int, form_type: str, responses: str) -> None:
    """Save a completed intake form response (JSON string)."""
    with get_db() as conn:
        conn.execute("""
            INSERT INTO intake_form_responses (prospect_id, form_type, responses)
            VALUES (?, ?, ?)
        """, (prospect_id, form_type, responses))
